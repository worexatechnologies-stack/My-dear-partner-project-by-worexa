import re

from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError as DjangoValidationError
from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from django.db.models import Count, Max, Q, Sum
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import permissions, serializers, status
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.views import APIView

from apps.accounts.models import (
    AccountType,
    Admin,
    AdminActivityLog,
    AdminPermission,
    AdminRole,
    AdminRolePermission,

    Member,
    MemberDocument,
    MemberPreference,
    MemberProfile,
    RoleCode,
    Staff,
    StaffActivityLog,
    SuperAdmin,
    SuperAdminActivityLog,
    Country,
    State,
    City,
    Branch,
    AdminAccessScope,
    StaffAccessScope,

    UserPermission,
    PermissionAuditLog,
)
from apps.accounts.serializers import MemberSerializer, administrative_account_payload
from apps.accounts.services import permanently_delete_member
from apps.accounts.verification_service import AccountVerificationService

from .api_utils import audit, bad_request, create_notification, create_ticket_attachment, notify, paginated_response
from .models import (
    PaymentOrder,
    PaymentTransaction,
    RefundRequest,
    RefundTransaction,
    RazorpayWebhookEvent,
    MembershipPurchase,
    BackupRecord,
    Complaint,
    ContactEnquiry,
    FAQ,
    MembershipPlan,
    Notification,
    Payment,
    MemberMembership,
    PlatformSetting,
    ProfileReport,
    ProfileVerificationAssignment,
    ProfileVerificationHistory,
    ProfileVerificationRequest,
    SupportCategory,
    SupportTicket,
    SupportTicketReply,
    TicketAssignment,
    TicketInternalNote,
    TicketStatusHistory,
)
from .responses import ApiResponse
from .serializers import (
    FAQSerializer,
    MemberTicketCreateSerializer,
    PaymentSerializer,
    ProfileVerificationSerializer,
    SupportTicketReplySerializer,
    SupportTicketSerializer,
    TicketInternalNoteSerializer,
    TicketReplyInputSerializer,
    administrative_summary,
    member_summary,
    MembershipPlanSerializer,
)


def notify_members_of_account_unavailability(*, member, action):
    """Make a member suspension/deletion visible to the rest of the platform.

    The notification deliberately does not identify the affected member.  An
    administrator's action must not expose a suspended or deleted member's
    personal information, while other active members still receive a clear
    explanation when an existing profile or conversation becomes unavailable.
    """
    action_label = "suspended" if action in {"deactivate", "suspend"} else "deleted"
    recipients = (
        Member.objects.filter(
            is_active=True,
            deleted_at__isnull=True,
            account_status=Member.AccountStatus.ACTIVE,
        )
        .exclude(pk=member.pk)
        .iterator(chunk_size=500)
    )
    for recipient in recipients:
        create_notification(
            recipient,
            type="MEMBER_ACCOUNT_UNAVAILABLE",
            title="Member account unavailable",
            body=(
                "A member account is no longer available because it has been "
                f"{action_label}."
            ),
            link_url="/notifications",
            priority="HIGH",
        )

class AccountScopePermission(permissions.BasePermission):
    message = 'This account type or role cannot access this endpoint.'

    def has_permission(self, request, view):
        user = request.user
        if not user or not user.is_authenticated or not user.is_active or user.deleted_at is not None:
            return False
        allowed = {str(value) for value in getattr(view, 'allowed_account_types', ())}
        if str(user.account_type) not in allowed:
            return False
        required = getattr(view, 'required_permission', None)
        if not required or str(user.account_type) == AccountType.SUPER_ADMIN:
            return True
        required = (required,) if isinstance(required, str) else tuple(required)
        return all(user.has_admin_permission(code) for code in required)

def apply_scope_filter(user, queryset, branch_path='branch'):
    if not user or not user.is_authenticated:
        return queryset.none()
    if user.account_type == AccountType.SUPER_ADMIN or getattr(user, 'is_superuser', False):
        return queryset
    
    scopes = []
    if user.account_type == AccountType.ADMIN:
        scopes = AdminAccessScope.objects.filter(account_id=user.id, is_active=True)
    elif user.account_type == AccountType.ADMIN:
        scopes = StaffAccessScope.objects.filter(account_id=user.id, is_active=True)
    branch_ids = []
    for sc in scopes:
        if sc.branch_id:
            branch_ids.append(sc.branch_id)
        elif sc.city_id:
            branch_ids.extend(Branch.objects.filter(city_id=sc.city_id).values_list('id', flat=True))
        elif sc.state_id:
            branch_ids.extend(Branch.objects.filter(city__state_id=sc.state_id).values_list('id', flat=True))
        elif sc.country_id:
            branch_ids.extend(Branch.objects.filter(city__state__country_id=sc.country_id).values_list('id', flat=True))
            
    if not branch_ids and getattr(user, 'branch_id', None):
        branch_ids = [user.branch_id]
        
    if not branch_ids:
        # No specific scope restrictions configured for this user - bypass filtering
        return queryset
        
    filter_kwargs = {f"{branch_path}__in": branch_ids}
    return queryset.filter(**filter_kwargs)

def check_object_scope(user, obj, branch_path='branch'):
    if not user or not user.is_authenticated:
        raise PermissionDenied('Authentication required.')
    if user.account_type == AccountType.SUPER_ADMIN or getattr(user, 'is_superuser', False):
        return
        
    obj_branch_id = obj
    for part in branch_path.split('__'):
        if obj_branch_id is None:
            break
        obj_branch_id = getattr(obj_branch_id, part, None)
        
    if obj_branch_id is not None and hasattr(obj_branch_id, 'id'):
        obj_branch_id = obj_branch_id.id
    elif obj_branch_id is not None:
        obj_branch_id = str(obj_branch_id)
        
    scopes = []
    if user.account_type == AccountType.ADMIN:
        scopes = AdminAccessScope.objects.filter(account_id=user.id, is_active=True)
    elif user.account_type == AccountType.ADMIN:
        scopes = StaffAccessScope.objects.filter(account_id=user.id, is_active=True)
    branch_ids = []
    for sc in scopes:
        if sc.branch_id:
            branch_ids.append(str(sc.branch_id))
        elif sc.city_id:
            branch_ids.extend(map(str, Branch.objects.filter(city_id=sc.city_id).values_list('id', flat=True)))
        elif sc.state_id:
            branch_ids.extend(map(str, Branch.objects.filter(city__state_id=sc.state_id).values_list('id', flat=True)))
        elif sc.country_id:
            branch_ids.extend(map(str, Branch.objects.filter(city__state__country_id=sc.country_id).values_list('id', flat=True)))
            
    if not branch_ids and getattr(user, 'branch_id', None):
        branch_ids = [str(user.branch_id)]
        
    if not branch_ids:
        # No specific scope restrictions configured for this user - bypass checks
        return
        
    if obj_branch_id and str(obj_branch_id) not in branch_ids:
        raise PermissionDenied('Access denied: The requested record belongs to another branch/scope.')

class ScopedAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated, AccountScopePermission)
    allowed_account_types = ()
    required_permission = None

class AdministrativeAccountCreateSerializer(serializers.Serializer):
    email = serializers.EmailField()
    mobile_number = serializers.CharField(required=False, allow_blank=True, allow_null=True)
    first_name = serializers.CharField(max_length=150)
    last_name = serializers.CharField(max_length=150, required=False, allow_blank=True)
    password = serializers.CharField(write_only=True)
    role = serializers.CharField(required=False, allow_blank=True, allow_null=True)
    photo = serializers.CharField(required=False, allow_blank=True, allow_null=True)
    bio = serializers.CharField(required=False, allow_blank=True, allow_null=True)
    employee_code = serializers.CharField(max_length=50, required=False, allow_blank=True, allow_null=True)
    manager_admin = serializers.CharField(max_length=50, required=False, allow_blank=True, allow_null=True)
    support_level = serializers.ChoiceField(
        choices=[],
        required=False,
    )
    specialization = serializers.ChoiceField(
        choices=[],
        required=False,
    )

    def validate_email(self, value):
        return value.lower()

    def validate_mobile_number(self, value):
        value = str(value or '').strip()
        if value and not re.fullmatch(r'^[6-9]\d{9}$', value):
            raise serializers.ValidationError('Enter a valid 10-digit Indian mobile number.')
        return value or None

    def validate_password(self, value):
        try:
            validate_password(value)
        except DjangoValidationError as exc:
            raise serializers.ValidationError(list(exc.messages)) from exc
        return value

    def validate(self, attrs):
        from apps.accounts.models import SuperAdmin, Admin, Staff, Member
        email = attrs.get('email', '').lower()
        mobile_number = attrs.get('mobile_number')
        employee_code = attrs.get('employee_code')

        # 1. Global Email uniqueness
        if email:
            if (SuperAdmin.objects.filter(email__iexact=email).exists() or
                Admin.objects.filter(email__iexact=email).exists() or
                Staff.objects.filter(email__iexact=email).exists() or
                Member.objects.filter(email__iexact=email).exists()):
                raise serializers.ValidationError({'email': 'This email is already registered in the system.'})

        # 2. Global Mobile uniqueness
        if mobile_number:
            mobile_number = str(mobile_number).strip()
            if (SuperAdmin.objects.filter(mobile_number=mobile_number).exists() or
                Admin.objects.filter(mobile_number=mobile_number).exists() or
                Staff.objects.filter(mobile_number=mobile_number).exists() or
                Member.objects.filter(mobile_number=mobile_number).exists()):
                raise serializers.ValidationError({'mobile_number': 'This mobile number is already registered in the system.'})

        # 3. Global Employee code uniqueness
        if employee_code:
            employee_code = str(employee_code).strip()
            if (Admin.objects.filter(employee_code__iexact=employee_code).exists() or
                Staff.objects.filter(employee_code__iexact=employee_code).exists()):
                raise serializers.ValidationError({'employee_code': 'This employee code is already in use.'})

        return attrs

def _create_administrative_account(*, actor, values, forced_role=None):
    role_code = str(forced_role or values.pop('role', ''))
    role = get_object_or_404(AdminRole, code=role_code)
    common = {
        'email': values['email'],
        'mobile_number': values.get('mobile_number'),
        'first_name': values['first_name'],
        'last_name': values.get('last_name', ''),
        'password': values['password'],
        'role': role,
        'photo': values.get('photo', ''),
        'bio': values.get('bio', ''),
        'is_email_verified': True,
    }
    
    from apps.accounts.models import Admin, Staff

    manager_admin = None
    manager_val = values.get('manager_admin')
    if manager_val:
        manager_admin = Admin.objects.filter(pk=manager_val).first()

    if role_code == RoleCode.ADMIN:
        if str(actor.account_type) != AccountType.SUPER_ADMIN:
            raise PermissionDenied('Only a Super Admin can create an Admin account.')
        if Admin.objects.filter(email__iexact=common['email']).exists():
            raise serializers.ValidationError({'email': ['An Admin with this email already exists.']})
        employee_code = values.get('employee_code') or f'ADM-{Admin.objects.count() + 1:05d}'
        return Admin.objects.create_user(
            created_by_super_admin=actor,
            employee_code=employee_code,
            **common
        )
        
    if str(actor.account_type) not in {AccountType.SUPER_ADMIN, AccountType.ADMIN}:
        raise PermissionDenied('Only a Super Admin or Admin can create operational accounts.')
    created_by_admin = actor if str(actor.account_type) == AccountType.ADMIN else None
    
    if role_code == RoleCode.ADMIN:
        if Staff.objects.filter(email__iexact=common['email']).exists():
            raise serializers.ValidationError({'email': ['A Staff account with this email already exists.']})
        employee_code = values.get('employee_code') or f'STF-{Staff.objects.count() + 1:05d}'
        return Staff.objects.create_user(
            employee_code=employee_code,
            manager_admin=manager_admin,
            created_by_admin=created_by_admin,
            **common,
        )
        
    raise serializers.ValidationError({'role': ['Unsupported account role.']})

def _get_administrative_account(account_id):
    for model in (Admin, Staff):
        account = model.objects.filter(pk=account_id).select_related('role').first()
        if account:
            return account
    return None

class SuperAdminAdminListCreateView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)
    required_permission = 'admins.manage'

    def get(self, request):
        return paginated_response(
            request,
            Admin.objects.filter(deleted_at__isnull=True).select_related('role'),
            _AdministrativeAccountSerializer,
        )

    @transaction.atomic
    def post(self, request):
        serializer = AdministrativeAccountCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        account = _create_administrative_account(
            actor=request.user,
            values=dict(serializer.validated_data),
            forced_role=RoleCode.ADMIN,
        )
        audit(
            request,
            request.user,
            action='ADMIN_CREATED',
            module='accounts',
            target_type='ADMIN',
            target_id=account.pk,
            new_data={'email': account.email},
        )
        return ApiResponse(
            data=administrative_account_payload(account),
            message='Admin account created.',
            status=status.HTTP_201_CREATED,
        )

class _AdministrativeAccountSerializer(serializers.Serializer):
    def to_representation(self, instance):
        return administrative_account_payload(instance)

class AdminAccountListCreateView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    def get(self, request):
        if not (
            str(request.user.account_type) == AccountType.SUPER_ADMIN
            or request.user.has_admin_permission('staff.view')
            or request.user.has_admin_permission('support_agents.view')
            or request.user.has_admin_permission('admins.manage_permissions')
        ):
            raise PermissionDenied()
        from django.db.models import Count, Q
        from django.utils import timezone
        role_filter = request.query_params.get('role', '').upper()
        search = request.query_params.get('search', '').strip()
        rows = []
        if role_filter in {'', RoleCode.ADMIN} and (str(request.user.account_type) == AccountType.SUPER_ADMIN or request.user.has_admin_permission('admins.manage_permissions')):
            queryset = Admin.objects.filter(deleted_at__isnull=True)
            if search:
                queryset = queryset.filter(
                    Q(email__icontains=search) | Q(first_name__icontains=search)
                    | Q(last_name__icontains=search) | Q(mobile_number__icontains=search)
                )
            rows.extend(queryset)
        if role_filter in {'', RoleCode.ADMIN}:
            queryset = Staff.objects.filter(deleted_at__isnull=True).annotate(
                assigned_count=Count('work_assignments', filter=Q(work_assignments__status='ASSIGNED')),
                in_progress_count=Count('work_assignments', filter=Q(work_assignments__status='IN_PROGRESS')),
                completed_count=Count('work_assignments', filter=Q(work_assignments__status='COMPLETED')),
                overdue_count=Count('work_assignments', filter=Q(work_assignments__status='ASSIGNED', work_assignments__due_at__lt=timezone.now()))
            )
            if search:
                queryset = queryset.filter(
                    Q(email__icontains=search) | Q(first_name__icontains=search)
                    | Q(last_name__icontains=search) | Q(mobile_number__icontains=search)
                )
            rows.extend(queryset)

        rows.sort(key=lambda row: row.created_at, reverse=True)
        return paginated_response(request, rows, _AdministrativeAccountSerializer)

    @transaction.atomic
    def post(self, request):
        serializer = AdministrativeAccountCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        values = dict(serializer.validated_data)
        role = str(values.get('role', ''))
        
        # Check permissions for creation
        if role == RoleCode.ADMIN:
            is_permitted = str(request.user.account_type) == AccountType.SUPER_ADMIN or request.user.has_admin_permission('admins.manage_permissions')
        else:
            permission = {
                RoleCode.ADMIN: 'staff.create',
            }.get(role)
            is_permitted = str(request.user.account_type) == AccountType.SUPER_ADMIN or (permission and request.user.has_admin_permission(permission))
            
        if not is_permitted:
            raise PermissionDenied('You cannot create this account type.')
            
        account = _create_administrative_account(actor=request.user, values=values)
        action = {
            RoleCode.ADMIN: 'ADMIN_CREATED',
            RoleCode.ADMIN: 'STAFF_CREATED',
        }[role]
        audit(
            request,
            request.user,
            action=action,
            module='accounts',
            target_type=role,
            target_id=account.pk,
            new_data={'email': account.email},
        )
        return ApiResponse(
            data=administrative_account_payload(account),
            message='Administrative account created.',
            status=status.HTTP_201_CREATED,
        )

class AdminAccountDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    def _account(self, request, account_id):
        account = _get_administrative_account(account_id)
        if not account:
            from rest_framework.exceptions import NotFound
            raise NotFound('Administrative account not found.')
            
        # Super Admin accounts can never be edited or restricted by non-SuperAdmins
        if str(account.account_type) == AccountType.SUPER_ADMIN:
            if str(request.user.account_type) != AccountType.SUPER_ADMIN:
                raise PermissionDenied('Super Admin accounts cannot be managed or modified by other roles.')
                
        if str(request.user.account_type) == AccountType.ADMIN and str(account.account_type) == AccountType.ADMIN:
            # Check if actor has explicit admins.manage_permissions
            if not request.user.has_admin_permission('admins.manage_permissions'):
                raise PermissionDenied('Admins cannot manage other Admin accounts unless explicitly permitted.')
                
        permission = {
            AccountType.ADMIN: 'admins.manage',
            AccountType.ADMIN: 'staff.manage',
        }[str(account.account_type)]
        if (
            str(request.user.account_type) != AccountType.SUPER_ADMIN
            and not request.user.has_admin_permission('admins.manage_permissions')
            and not request.user.has_admin_permission(permission)
        ):
            raise PermissionDenied('You cannot manage this account type.')
        return account

    def get(self, request, account_id):
        return ApiResponse(data=administrative_account_payload(self._account(request, account_id)))

    @transaction.atomic
    def patch(self, request, account_id):
        account = self._account(request, account_id)
        before = {'is_active': account.is_active, 'email': account.email}
        update_fields = []
        for field in ('first_name', 'last_name', 'mobile_number', 'is_active', 'photo', 'bio'):
            if field in request.data:
                setattr(account, field, request.data[field])
                update_fields.append(field)
                
        action = request.data.get('action')
        if action in {'activate', 'deactivate'}:
            should_be_active = action == 'activate'
            if account.is_active != should_be_active:
                account.is_active = should_be_active
                update_fields.append('is_active')
                if not should_be_active:
                    account.token_version += 1
                    update_fields.append('token_version')
        elif action == 'reset_password':
            if not request.data.get('new_password'):
                return bad_request('new_password is required.')
            try:
                validate_password(request.data['new_password'], user=account)
            except DjangoValidationError as exc:
                return bad_request('Choose a stronger password.', errors={'new_password': exc.messages})
            account.set_password(request.data['new_password'])
            account.password_changed_at = timezone.now()
            account.token_version += 1
            update_fields.extend(('password', 'password_changed_at', 'token_version'))
        elif action:
            return bad_request('Unsupported account action.')

        # Database-backed updates
        from apps.accounts.models import Admin, SuperAdmin, Staff
        
        if 'employee_code' in request.data and hasattr(account, 'employee_code'):
            account.employee_code = request.data['employee_code']
            update_fields.append('employee_code')

        if 'support_level' in request.data and hasattr(account, 'support_level'):
            account.support_level = request.data['support_level']
            update_fields.append('support_level')

        if 'specialization' in request.data and hasattr(account, 'specialization'):
            account.specialization = request.data['specialization']
            update_fields.append('specialization')

        if 'manager_admin' in request.data and hasattr(account, 'manager_admin'):
            manager_val = request.data['manager_admin']
            if manager_val:
                account.manager_admin = Admin.objects.filter(pk=manager_val).first()
            else:
                account.manager_admin = None
            update_fields.append('manager_admin')

        if request.data.get('role') and str(request.data['role']) != str(account.account_type):
            return bad_request('Changing account type requires creating the account in its own table.')
            
        if update_fields:
            # Using basic save instead of update_fields since we might have fields from multiple child models/bases
            account.save()
            
        audit(
            request,
            request.user,
            action='ACCOUNT_UPDATED',
            module='accounts',
            target_type=str(account.account_type),
            target_id=account.pk,
            old_data=before,
            new_data={'is_active': account.is_active, 'email': account.email},
        )
        return ApiResponse(data=administrative_account_payload(account), message='Account updated.')

    def delete(self, request, account_id):
        account = self._account(request, account_id)
        
        # Deletion safety rules
        from apps.accounts.models import SuperAdmin, AccountType
        if str(account.account_type) == AccountType.SUPER_ADMIN:
            if SuperAdmin.objects.filter(deleted_at__isnull=True).count() <= 1:
                return bad_request('Cannot delete the final Super Admin.')
        
        if str(account.pk) == str(request.user.pk):
            return bad_request('Cannot delete your own currently logged-in account.')
            
        from apps.core.models import WorkAssignment
        if str(account.account_type) == 'ADMIN':
            active_tasks = WorkAssignment.objects.filter(
                assigned_to_staff_id=account.pk, 
                status__in=['ASSIGNED', 'IN_PROGRESS', 'WAITING', 'ESCALATED']
            ).exists()
            if active_tasks:
                return bad_request('Cannot delete a Staff member with active assigned work.')

        account.delete()
        audit(
            request,
            request.user,
            action='ACCOUNT_DELETED',
            module='accounts',
            target_type=str(account.account_type),
            target_id=account_id,
        )
        return ApiResponse(message='Admin account permanently deleted.')

class AdminAccountApprovalsView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    def get(self, request, account_id):
        from django.db.models import Q
        from apps.core.models import ProfileVerificationRequest
        from apps.profiles.models import ProfilePhoto
        from apps.accounts.models import MemberDocument

        profiles = ProfileVerificationRequest.objects.filter(
            status=ProfileVerificationRequest.Status.APPROVED,
            verification_type=ProfileVerificationRequest.VerificationType.FULL_PROFILE,
        ).filter(
            Q(reviewed_by_admin_id=account_id) | Q(reviewed_by_super_admin_id=account_id) | Q(reviewed_by_staff_id=account_id)
        ).select_related('member').order_by('-reviewed_at')[:50]

        photos = ProfilePhoto.objects.filter(
            status=ProfilePhoto.Status.APPROVED,
        ).filter(
            Q(verified_by_admin_id=account_id) | Q(verified_by_staff_id=account_id) | Q(verified_by_super_admin_id=account_id)
        ).select_related('user').order_by('-verified_at')[:50]

        documents = MemberDocument.objects.filter(
            status=MemberDocument.Status.APPROVED,
            reviewed_by_id=account_id,
        ).select_related('member').order_by('-reviewed_at')[:50]

        from apps.accounts.models import AdminActivityLog
        profile_items = [{
            'id': str(p.id),
            'member_id': str(p.member_id),
            'member_name': p.member.get_full_name() if p.member else 'Member',
            'member_email': p.member.email if p.member else '',
            'approved_at': p.reviewed_at.isoformat() if p.reviewed_at else None,
        } for p in profiles]

        if not profile_items:
            prof_logs = AdminActivityLog.objects.filter(
                actor_id=account_id
            ).filter(
                Q(action__icontains='APPROVE_PROFILE') | Q(action='MEMBER_APPROVE_PROFILE') | Q(action='PROFILE_APPROVED')
            ).order_by('-created_at')[:50]

            for log in prof_logs:
                profile_items.append({
                    'id': str(log.id),
                    'member_id': str(log.target_id) if log.target_id else '',
                    'member_name': log.target_id_str or f"Member #{str(log.target_id)[:8]}" if log.target_id else 'Member',
                    'member_email': '',
                    'approved_at': log.created_at.isoformat(),
                })

        return ApiResponse(data={
            'profiles_approved': {
                'count': len(profile_items),
                'items': profile_items,
            },
            'photos_approved': {
                'count': photos.count(),
                'items': [{
                    'id': str(p.id),
                    'member_id': str(p.user_id),
                    'member_name': p.user.get_full_name() if p.user else '',
                    'approved_at': p.verified_at.isoformat() if p.verified_at else None,
                } for p in photos],
            },
            'documents_approved': {
                'count': documents.count(),
                'items': [{
                    'id': str(d.id),
                    'member_id': str(d.member_id),
                    'member_name': d.member.get_full_name() if d.member else '',
                    'document_type': d.document_type,
                    'approved_at': d.reviewed_at.isoformat() if d.reviewed_at else None,
                } for d in documents],
            },
        })


def _admin_role_payload(role):
    return {
        'id': str(role.pk),
        'code': role.code,
        'name': role.name,
        'description': role.description,
        'is_system_role': role.is_system_role,
        'permissions': [
            row.permission.code for row in role.role_permissions.all() if row.is_allowed
        ],
    }

class AdminRolePermissionView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)
    required_permission = 'roles.manage'

    def get(self, _request, role_id=None):
        permissions_data = list(
            AdminPermission.objects.values('id', 'code', 'name', 'module', 'description')
        )
        roles = AdminRole.objects.prefetch_related('role_permissions__permission')
        if role_id:
            roles = roles.filter(pk=role_id)
        role_data = [_admin_role_payload(role) for role in roles]
        return ApiResponse(data={'roles': role_data, 'permissions': permissions_data})

    @transaction.atomic
    def put(self, request, role_id=None):
        role = get_object_or_404(AdminRole, pk=role_id)
        if role.code == RoleCode.SUPER_ADMIN:
            return bad_request('Super Admin always has every permission and cannot be restricted.')
        requested_codes = set(request.data.get('permissions', []))
        known = set(AdminPermission.objects.values_list('code', flat=True))
        unknown = requested_codes - known
        if unknown:
            return bad_request('Unknown permission codes.', errors={'permissions': sorted(unknown)})
        before = list(
            role.role_permissions.filter(is_allowed=True).values_list('permission__code', flat=True)
        )
        for permission in AdminPermission.objects.all():
            AdminRolePermission.objects.update_or_create(
                role=role,
                permission=permission,
                defaults={'is_allowed': permission.code in requested_codes},
            )
        audit(
            request,
            request.user,
            action='ROLE_PERMISSIONS_UPDATED',
            module='roles',
            target_type='ADMIN_ROLE',
            target_id=role.pk,
            old_data={'permissions': before},
            new_data={'permissions': sorted(requested_codes)},
        )
        role = AdminRole.objects.prefetch_related('role_permissions__permission').get(pk=role.pk)
        return ApiResponse(data=_admin_role_payload(role), message='Role permissions updated.')

class AdminUserPermissionsView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    def _get_target_user(self, user_id):
        from apps.accounts.models import Admin, Staff
        for model in (Admin, Staff):
            user = model.objects.filter(pk=user_id, deleted_at__isnull=True).first()
            if user:
                return user
        return None

    def get(self, request, user_id):
        target_user = self._get_target_user(user_id)
        if not target_user:
            return ApiResponse(message='User not found.', status=status.HTTP_404_NOT_FOUND)
        
        actor = request.user
        
        # Guard access:
        # Super Admin can manage anyone.
        # Admins can manage Staff and Support agents.
        # Admins can manage other Admins ONLY IF they have 'admins.manage_permissions'.
        if str(actor.account_type) == AccountType.ADMIN:
            if str(target_user.account_type) == AccountType.ADMIN:
                if not actor.has_admin_permission('admins.manage_permissions'):
                    raise PermissionDenied('You do not have permission to manage other Admins.')
            elif str(target_user.account_type) == AccountType.SUPER_ADMIN:
                raise PermissionDenied('Super Admin accounts cannot be managed.')
        elif str(actor.account_type) == AccountType.SUPER_ADMIN:
            if str(target_user.account_type) == AccountType.SUPER_ADMIN:
                # Super Admin permissions are immutable
                return ApiResponse(message='Super Admin permissions are immutable.', status=status.HTTP_400_BAD_REQUEST)
        
        # Load role permissions for target user
        role_permissions = set()
        if target_user.role:
            role_permissions = set(
                target_user.role.role_permissions.filter(is_allowed=True).values_list('permission__code', flat=True)
            )
            
        # Load overrides for target user
        from django.contrib.contenttypes.models import ContentType
        from apps.accounts.models import UserPermission, AdminPermission
        
        ct = ContentType.objects.get_for_model(target_user.__class__)
        overrides = {
            up.permission.code: up.is_allowed
            for up in UserPermission.objects.filter(user_content_type=ct, user_object_id=target_user.id).select_related('permission')
        }
        
        # Fetch all available permissions
        permissions = list(
            AdminPermission.objects.all().order_by('module', 'code')
        )
        
        data_perms = []
        for perm in permissions:
            code = perm.code
            is_inherited = code in role_permissions
            is_overridden = code in overrides
            override_val = overrides.get(code, None)
            
            # Effective permission
            is_allowed = override_val if is_overridden else is_inherited
            
            # Can actor grant/modify this permission?
            can_grant = actor.has_admin_permission(code)
            if code == 'admins.manage_permissions' and str(actor.account_type) != AccountType.SUPER_ADMIN:
                can_grant = False
                
            data_perms.append({
                'code': code,
                'name': perm.name,
                'module': perm.module,
                'description': perm.description,
                'is_inherited': is_inherited,
                'is_overridden': is_overridden,
                'is_allowed': is_allowed,
                'can_grant': can_grant,
            })
            
        # Target user serialization
        user_info = {
            'id': str(target_user.id),
            'first_name': target_user.first_name,
            'last_name': target_user.last_name,
            'full_name': target_user.get_full_name(),
            'email': target_user.email,
            'role': str(target_user.account_type),
            'role_display': target_user.admin_role_display,
            'is_active': target_user.is_active,
        }
        
        return ApiResponse(data={
            'user': user_info,
            'permissions': data_perms,
        })

    @transaction.atomic
    def post(self, request, user_id):
        target_user = self._get_target_user(user_id)
        if not target_user:
            return ApiResponse(message='User not found.', status=status.HTTP_404_NOT_FOUND)
        
        actor = request.user
        
        # Guard access:
        if str(actor.account_type) == AccountType.ADMIN:
            if str(target_user.account_type) == AccountType.ADMIN:
                if not actor.has_admin_permission('admins.manage_permissions'):
                    raise PermissionDenied('You do not have permission to manage other Admins.')
            elif str(target_user.account_type) == AccountType.SUPER_ADMIN:
                raise PermissionDenied('Super Admin accounts cannot be managed.')
        elif str(actor.account_type) == AccountType.SUPER_ADMIN:
            if str(target_user.account_type) == AccountType.SUPER_ADMIN:
                return ApiResponse(message='Super Admin permissions are immutable.', status=status.HTTP_400_BAD_REQUEST)
        
        requested_overrides = request.data.get('permissions', []) # expects [{"code": "users.view", "is_allowed": true/false/null}, ...]
        
        from django.contrib.contenttypes.models import ContentType
        from apps.accounts.models import UserPermission, PermissionAuditLog, AdminPermission
        
        ct = ContentType.objects.get_for_model(target_user.__class__)
        actor_ct = ContentType.objects.get_for_model(actor.__class__)
        
        # Load previous overrides
        prev_overrides = {
            up.permission.code: up.is_allowed
            for up in UserPermission.objects.filter(user_content_type=ct, user_object_id=target_user.id).select_related('permission')
        }
        
        # Validate and apply overrides
        new_overrides = dict(prev_overrides)
        
        for override in requested_overrides:
            code = override.get('code')
            val = override.get('is_allowed') # True, False, or None
            
            perm = AdminPermission.objects.filter(code=code).first()
            if not perm:
                return bad_request(f'Unknown permission code: {code}')
                
            # Check privilege escalation
            # Actor must possess this permission to modify it
            if not actor.has_admin_permission(code):
                return bad_request(f'Privilege escalation warning: You do not have permission "{code}" to grant or deny it.')
                
            # Only Super Admin can grant admins.manage_permissions
            if code == 'admins.manage_permissions' and str(actor.account_type) != AccountType.SUPER_ADMIN:
                return bad_request('Only the Super Admin can modify admins.manage_permissions.')
                
            if val is None:
                # Remove override
                UserPermission.objects.filter(user_content_type=ct, user_object_id=target_user.id, permission=perm).delete()
                new_overrides.pop(code, None)
            else:
                # Save override
                UserPermission.objects.update_or_create(
                    user_content_type=ct,
                    user_object_id=target_user.id,
                    permission=perm,
                    defaults={'is_allowed': val, 'assigned_by_content_type': actor_ct, 'assigned_by_object_id': actor.id}
                )
                new_overrides[code] = val
                
        # Record Permission Audit Log
        if prev_overrides != new_overrides:
            PermissionAuditLog.objects.create(
                user_content_type=ct,
                user_object_id=target_user.id,
                actor_content_type=actor_ct,
                actor_object_id=actor.id,
                previous_permissions=prev_overrides,
                new_permissions=new_overrides,
                ip_address=self.get_client_ip(request)
            )
            
            # Force log out by bumping token version to revoke existing sessions if permissions changed
            target_user.token_version += 1
            target_user.save(update_fields=('token_version', 'updated_at'))
            
        return ApiResponse(message='User permissions updated successfully.')

    def get_client_ip(self, request):
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0].strip()
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

class _PermissionAuditLogSerializer(serializers.ModelSerializer):
    actor_name = serializers.SerializerMethodField()
    user_name = serializers.SerializerMethodField()
    user_email = serializers.SerializerMethodField()
    user_role = serializers.SerializerMethodField()

    class Meta:
        model = PermissionAuditLog
        fields = (
            'id', 'user_name', 'user_email', 'user_role', 'actor_name', 
            'previous_permissions', 'new_permissions', 'ip_address', 'created_at'
        )

    def get_actor_name(self, obj):
        return obj.actor.get_full_name() if obj.actor else 'System'

    def get_user_name(self, obj):
        return obj.user.get_full_name() if obj.user else 'Unknown'

    def get_user_email(self, obj):
        return obj.user.email if obj.user else ''

    def get_user_role(self, obj):
        return str(obj.user.account_type) if obj.user else ''

class AdminPermissionAuditLogListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'audit_logs.view'

    def get(self, request):
        logs = PermissionAuditLog.objects.all().order_by('-created_at')
        return paginated_response(request, logs, _PermissionAuditLogSerializer)

class AdminDashboardView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'dashboard.view'

    def _date_range(self, range_param, now):
        from datetime import timedelta
        if range_param == '7d':
            return now - timedelta(days=7)
        if range_param == '30d':
            return now - timedelta(days=30)
        if range_param == '90d':
            return now - timedelta(days=90)
        if range_param == '1y':
            return now - timedelta(days=365)
        return None  # all time

    def get(self, request):
        from django.db.models import Sum, Count
        from django.utils import timezone
        from datetime import timedelta

        is_super = str(request.user.account_type) == AccountType.SUPER_ADMIN

        # Date helpers
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        range_start = self._date_range(request.query_params.get('range', '30d'), now)

        # Member counts
        members_qs = Member.objects.filter(deleted_at__isnull=True)
        total_users = members_qs.count()
        active_users = members_qs.filter(is_active=True).count()
        suspended_users = members_qs.filter(is_active=False).count()
        new_today = members_qs.filter(created_at__gte=today_start).count()
        new_this_month = members_qs.filter(created_at__gte=month_start).count()
        if range_start:
            members_qs = members_qs.filter(created_at__gte=range_start)
        verified_users = members_qs.filter(profile_status=Member.ProfileStatus.APPROVED).count()
        male_profiles = members_qs.filter(gender='Male').count()
        female_profiles = members_qs.filter(gender='Female').count()
        premium_users = members_qs.filter(is_premium=True).count()

        # Verification queues
        pending_profile_approvals = ProfileVerificationRequest.objects.filter(
            verification_type=ProfileVerificationRequest.VerificationType.FULL_PROFILE,
            status__in=(
                ProfileVerificationRequest.Status.PENDING_REVIEW,
            )
        ).count()
        pending_photo_approvals = members_qs.filter(photo_status=Member.VerificationStatus.PENDING_REVIEW).count()

        # Document verification
        try:
            from apps.accounts.models import MemberDocument
            pending_document_verification = MemberDocument.objects.filter(status='PENDING').count()
        except Exception:
            pending_document_verification = 0

        # Memberships
        try:
            from apps.core.models import MemberMembership
            active_memberships = MemberMembership.objects.filter(
                is_active=True, expires_at__gte=now
            ).count()
            expired_memberships = MemberMembership.objects.filter(
                is_active=False
            ).count()
        except Exception:
            active_memberships = 0
            expired_memberships = 0

        # Payments
        from django.db.models import Q as DQ
        total_revenue_val = PaymentTransaction.objects.filter(
            status='captured'
        ).aggregate(total=Sum('amount'))['total'] or 0
        revenue_this_month_val = PaymentTransaction.objects.filter(
            status='captured', created_at__gte=month_start
        ).aggregate(total=Sum('amount'))['total'] or 0
        total_revenue = str(total_revenue_val)
        revenue_this_month = str(revenue_this_month_val)
        pending_payments = PaymentOrder.objects.filter(status='created').count()
        successful_payments = PaymentTransaction.objects.filter(status='captured').count()
        failed_payments = PaymentTransaction.objects.filter(status='failed').count()

        # Support tickets
        pending_tickets = SupportTicket.objects.filter(status=SupportTicket.Status.UNASSIGNED).count()
        assigned_tickets = SupportTicket.objects.filter(status=SupportTicket.Status.ASSIGNED).count()
        open_tickets = SupportTicket.objects.exclude(
            status__in=(SupportTicket.Status.RESOLVED, SupportTicket.Status.CLOSED)
        ).count()

        # Complaints & reports
        open_complaints = Complaint.objects.filter(status='OPEN').count()
        reported_profiles = ProfileReport.objects.filter(status='OPEN').count()

        stats = {
            'total_users': total_users,
            'active_users': active_users,
            'suspended_users': suspended_users,
            'new_today': new_today,
            'new_this_month': new_this_month,
            'verified_users': verified_users,
            'male_profiles': male_profiles,
            'female_profiles': female_profiles,
            'premium_users': premium_users,
            'pending_profile_approvals': pending_profile_approvals,
            'pending_photo_approvals': pending_photo_approvals,
            'pending_document_verification': pending_document_verification,
            'pending_verification': pending_profile_approvals,  # legacy alias
            'active_memberships': active_memberships,
            'expired_memberships': expired_memberships,
            'total_revenue': total_revenue,
            'revenue': total_revenue,  # legacy alias
            'revenue_this_month': revenue_this_month,
            'pending_payments': pending_payments,
            'successful_payments': successful_payments,
            'failed_payments': failed_payments,
            'pending_tickets': pending_tickets,
            'assigned_tickets': assigned_tickets,
            'open_tickets': open_tickets,
            'open_complaints': open_complaints,
            'reported_profiles': reported_profiles,
            'escalated_complaints': Complaint.objects.filter(status='ESCALATED').count() if hasattr(Complaint, 'objects') else 0,
        }

        recent_member_qs = Member.objects.filter(deleted_at__isnull=True)
        if range_start:
            recent_member_qs = recent_member_qs.filter(created_at__gte=range_start)
        from django.db.models import Prefetch
        from apps.core.models import MemberMembership, MembershipRequest
        recent_member_qs = recent_member_qs.prefetch_related(
            Prefetch('memberships',
                queryset=MemberMembership.objects.filter(is_active=True).select_related('plan'),
                to_attr='_prefetched_active_memberships'),
            Prefetch('memberships',
                queryset=MemberMembership.objects.filter(status='PENDING_VERIFICATION').select_related('plan'),
                to_attr='_prefetched_pending_memberships'),
            Prefetch('membership_requests',
                queryset=MembershipRequest.objects.filter(status='pending').select_related('selected_plan'),
                to_attr='_prefetched_pending_requests'),
        )
        recent_members = recent_member_qs.order_by('-created_at')[:5]
        recent_tickets = SupportTicket.objects.select_related('category').order_by('-created_at')[:5]
        recent_payments = PaymentTransaction.objects.select_related('payment_order__membership_plan', 'user').order_by('-created_at')[:5]

        return ApiResponse(data={
            'role': request.user.admin_role_code,
            'role_display': request.user.admin_role_display,
            'permissions': sorted(request.user.get_effective_admin_permissions()),
            'stats': stats,
            'charts': {'registrations': [], 'revenue': [], 'memberships': []},
            'recent_users': MemberSerializer(recent_members, many=True, context={'request': request}).data,
            'recent_tickets': SupportTicketSerializer(recent_tickets, many=True).data,
            'recent_payments': PaymentSerializer(recent_payments, many=True).data,
            'recent_activity': [],
            'monthly_signups': [],
            'membership_distribution': [],
            'content': {
                'faqs': FAQ.objects.count(),
            },
        })

class AdminUserListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'members.view'

    ORDERING_FIELDS = {
        'date_joined': 'created_at',
        '-date_joined': '-created_at',
        'first_name': 'first_name',
        '-first_name': '-first_name',
        'created_at': 'created_at',
        '-created_at': '-created_at',
    }

    def get(self, request):
        from django.db.models import Prefetch
        from apps.core.models import MemberMembership, MembershipRequest, Interest, ProfileUnlock
        from django.utils import timezone
        import zoneinfo

        kolkata_tz = zoneinfo.ZoneInfo("Asia/Kolkata")
        today = timezone.now().astimezone(kolkata_tz).date()

        queryset = Member.objects.filter(deleted_at__isnull=True).select_related('profile', 'preferences')
        queryset = queryset.prefetch_related(
            Prefetch('memberships',
                queryset=MemberMembership.objects.filter(is_active=True).select_related('plan'),
                to_attr='_prefetched_active_memberships'),
            Prefetch('memberships',
                queryset=MemberMembership.objects.filter(status='PENDING_VERIFICATION').select_related('plan'),
                to_attr='_prefetched_pending_memberships'),
            Prefetch('membership_requests',
                queryset=MembershipRequest.objects.filter(status='pending').select_related('selected_plan'),
                to_attr='_prefetched_pending_requests'),
            Prefetch('profile_unlocks_made',
                queryset=ProfileUnlock.objects.filter(usage_date=today),
                to_attr='_prefetched_today_unlocks'),
            Prefetch('sent_interests',
                queryset=Interest.objects.filter(created_at__date=today),
                to_attr='_prefetched_today_interests'),
        )
        queryset = apply_scope_filter(request.user, queryset, branch_path='branch')
        search = request.query_params.get('search', '').strip()
        if search:
            try:
                import uuid
                search_uuid = uuid.UUID(search)
                queryset = queryset.filter(id=search_uuid)
            except ValueError:
                queryset = queryset.filter(
                    Q(email__icontains=search) | Q(first_name__icontains=search) | Q(last_name__icontains=search)
                )
        requested_status = request.query_params.get('status')
        if requested_status == 'deleted':
            queryset = Member.objects.filter(deleted_at__isnull=False).select_related('profile', 'preferences')
        elif requested_status:
            queryset = queryset.filter(profile_status=requested_status)
        ordering = request.query_params.get('ordering', '-date_joined')
        order_field = self.ORDERING_FIELDS.get(ordering)
        if order_field:
            queryset = queryset.order_by(order_field)
        else:
            queryset = queryset.order_by('-created_at')
        return paginated_response(request, queryset, MemberSerializer, context={'today': today})

class AdminUserActionView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    # Member operations have distinct privileges.  Keep this mapping at the
    # API boundary so a UI bug cannot let a user with, for example, only
    # suspension access delete a member.
    action_permissions = {
        'approve_profile': 'members.manage',
        'verify': 'members.manage',
        'unverify': 'members.manage',
        'reject_profile': 'members.manage',
        'approve_photo': 'members.manage',
        'reject_photo': 'members.manage',
        'verify_document': 'members.manage',
        'reject_document': 'members.manage',
        'activate': 'members.suspend',
        'reactivate': 'members.suspend',
        'deactivate': 'members.suspend',
        'suspend': 'members.suspend',
        'soft_delete': 'members.delete',
        'restore': 'members.delete',
        'permanent_delete': 'members.delete',
    }

    @staticmethod
    def _require_permission(request, permission):
        if (
            str(request.user.account_type) != AccountType.SUPER_ADMIN
            and not request.user.has_admin_permission(permission)
        ):
            raise PermissionDenied(f'{permission} permission is required.')

    def get(self, request, user_id):
        self._require_permission(request, 'members.view')
        queryset = Member.objects.select_related('profile', 'preferences')
        member = get_object_or_404(queryset, pk=user_id)
        check_object_scope(request.user, member, branch_path='branch')
        activity_model = SuperAdminActivityLog if str(request.user.account_type) == AccountType.SUPER_ADMIN else AdminActivityLog
        activity = activity_model.objects.filter(target_type='MEMBER', target_id=str(member.pk)).values(
            'id', 'action', 'module', 'description', 'created_at', 'old_data', 'new_data'
        )[:50]
        verifications = ProfileVerificationRequest.objects.filter(member=member).values(
            'id', 'verification_type', 'status', 'submitted_at', 'reviewed_at', 'rejection_reason'
        )
        memberships = MemberMembership.objects.filter(member=member).select_related('plan').values(
            'id', 'status', 'is_active', 'start_date', 'end_date', 'plan__name', 'plan__slug'
        )
        documents = member.documents.values('id', 'document_type', 'status', 'uploaded_at', 'rejection_reason', 'reviewed_at')
        from apps.profiles.models import ProfilePhoto
        photos = ProfilePhoto.objects.without_binary().filter(user=member).values(
            'id', 'status', 'is_primary', 'display_order', 'original_filename',
            'created_at', 'verified_at', 'rejection_reason'
        )
        return ApiResponse(data={
            'member': MemberSerializer(member, context={'request': request}).data,
            'photos': list(photos),
            'verifications': list(verifications),
            'documents': list(documents),
            'memberships': list(memberships),
            'activity': list(activity),
        })

    @transaction.atomic
    def patch(self, request, user_id):
        member = get_object_or_404(Member.objects.select_for_update(), pk=user_id)
        check_object_scope(request.user, member, branch_path='branch')
        action = request.data.get('action')
        permission = self.action_permissions.get(action)
        if permission is None:
            return bad_request('Unsupported member action.')
        self._require_permission(request, permission)
        reason = str(request.data.get('reason', '')).strip()
        before = {
            'is_active': member.is_active,
            'profile_status': member.profile_status,
            'photo_status': member.photo_status,
            'document_status': member.document_status,
        }
        if action == 'approve_profile':
            # Delegate to the verification service so the approval is applied
            # consistently (profile_status, reviewed_at, and the linked
            # ProfileVerificationRequest are all updated in one place).
            AccountVerificationService.approve_profile(member, request.user, reason)
        elif action == 'verify':
            # An administrator's explicit verification must satisfy both
            # contact checks enforced by membership checkout.
            member.is_email_verified = True
            member.is_mobile_verified = True
        elif action == 'unverify':
            member.is_email_verified = False
            member.is_mobile_verified = False
        elif action == 'reject_profile':
            if not reason:
                return bad_request('A rejection reason is required.')
            AccountVerificationService.reject_profile(member, request.user, reason)
        elif action == 'approve_photo':
            member.photo_status = Member.VerificationStatus.APPROVED
            from apps.profiles.models import ProfilePhoto
            from apps.profiles.services.photo_management import _reviewer_fields
            reviewer_args = _reviewer_fields(request.user)
            ProfilePhoto.objects.filter(user=member).exclude(
                status=ProfilePhoto.Status.APPROVED
            ).update(
                status=ProfilePhoto.Status.APPROVED,
                verified_at=timezone.now(),
                **reviewer_args,
            )
        elif action == 'reject_photo':
            if not reason:
                return bad_request('A rejection reason is required.')
            member.photo_status = Member.VerificationStatus.REJECTED
            member.photo_rejection_reason = reason
            from apps.profiles.models import ProfilePhoto
            ProfilePhoto.objects.filter(user=member, is_primary=True).exclude(
                status=ProfilePhoto.Status.REJECTED
            ).update(
                status=ProfilePhoto.Status.REJECTED,
                verified_at=timezone.now(),
                rejection_reason=reason,
            )
        elif action == 'verify_document':
            document_id = request.data.get('document_id')
            documents = member.documents.filter(status=MemberDocument.Status.PENDING)
            if document_id:
                documents = documents.filter(pk=document_id)
            document = documents.order_by('-uploaded_at').first()
            if document is None:
                return bad_request('No pending KYC document is available to approve.')
            document.status = MemberDocument.Status.APPROVED
            document.rejection_reason = ''
            document.reviewed_at = timezone.now()
            document.reviewed_by_id = request.user.pk
            document.save(update_fields=('status', 'rejection_reason', 'reviewed_at', 'reviewed_by_id'))
            member.document_status = Member.VerificationStatus.APPROVED
        elif action == 'reject_document':
            if not reason:
                return bad_request('A rejection reason is required.')
            document_id = request.data.get('document_id')
            documents = member.documents.filter(status=MemberDocument.Status.PENDING)
            if document_id:
                documents = documents.filter(pk=document_id)
            document = documents.order_by('-uploaded_at').first()
            if document is None:
                return bad_request('No pending KYC document is available to reject.')
            document.status = MemberDocument.Status.REJECTED
            document.rejection_reason = reason
            document.reviewed_at = timezone.now()
            document.reviewed_by_id = request.user.pk
            document.save(update_fields=('status', 'rejection_reason', 'reviewed_at', 'reviewed_by_id'))
            member.document_status = Member.VerificationStatus.REJECTED
        elif action in {'activate', 'reactivate', 'restore'}:
            member.is_active = True
            member.deleted_at = None
            member.account_status = Member.AccountStatus.ACTIVE
            if member.profile_status == 'SUSPENDED':
                member.profile_status = Member.VerificationStatus.APPROVED
        elif action in {'deactivate', 'suspend'}:
            member.is_active = False
            member.account_status = Member.AccountStatus.SUSPENDED
            member.profile_status = 'SUSPENDED'
            member.token_version += 1
        elif action == 'soft_delete':
            member.is_active = False
            member.account_status = Member.AccountStatus.DELETED
            member.deleted_at = timezone.now()
            member.token_version += 1
        elif action == 'permanent_delete':
            if str(request.user.account_type) != AccountType.SUPER_ADMIN:
                raise PermissionDenied('Only a Super Admin can permanently delete a member.')
            target_id = member.pk
            deletion_result = permanently_delete_member(member=member, actor=request.user)
            audit(
                request, request.user, action='MEMBER_PERMANENTLY_DELETED', module='members',
                target_type='MEMBER', target_id=target_id,
                old_data={**before, **deletion_result.audit_context()},
            )
            return ApiResponse(message='Member permanently deleted.')
        member.save()
        after = {
            'is_active': member.is_active,
            'profile_status': member.profile_status,
            'photo_status': member.photo_status,
            'document_status': member.document_status,
        }
        audit(
            request, request.user, action=f'MEMBER_{str(action).upper()}', module='members',
            target_type='MEMBER', target_id=member.pk, old_data=before, new_data=after,
        )
        if action in {'deactivate', 'suspend', 'soft_delete'}:
            notify_members_of_account_unavailability(member=member, action=action)
        return ApiResponse(data=MemberSerializer(member, context={'request': request}).data)

class AdminGrantMembershipView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'members.manage'

    @transaction.atomic
    def post(self, request, user_id):
        member = get_object_or_404(Member.objects.select_for_update(), pk=user_id)
        check_object_scope(request.user, member, branch_path='branch')

        plan_slug = str(request.data.get('plan_slug') or '').strip().lower()
        if not plan_slug:
            return bad_request('Please select an active membership plan.')
        try:
            duration_months = int(request.data.get('duration_months', 12))
        except (TypeError, ValueError):
            return bad_request('Choose a valid membership duration.')
        if duration_months not in (1, 3, 6, 12, 24, 36):
            return bad_request('Choose a membership duration between 1 and 36 months.')
        
        from apps.core.services.membership_service import MembershipService
        success, message, membership = MembershipService.grant_custom_membership(
            member=member,
            plan_slug=plan_slug,
            duration_months=duration_months,
            actor=request.user,
            reason=f'Admin grant ({duration_months} months)',
        )
        if not success:
            return bad_request(message)
        
        return ApiResponse(message=message, data={
            'plan_name': membership.plan.name,
            'duration_months': duration_months,
            'expires_at': membership.end_date.isoformat() if membership.end_date else None,
        })

    @transaction.atomic
    def put(self, request, user_id):
        self._require_permission(request, 'members.manage')
        member = get_object_or_404(Member.objects.select_for_update(), pk=user_id)
        check_object_scope(request.user, member, branch_path='branch')
        before = MemberSerializer(member, context={'request': request}).data

        allowed_fields = {
            'first_name', 'last_name', 'gender', 'date_of_birth',
            'is_active', 'is_premium', 'is_email_verified', 'is_mobile_verified',
        }
        profile_fields = {
            'marital_status', 'height', 'weight', 'blood_group', 'complexion',
            'religion', 'mother_tongue', 'caste', 'sub_caste', 'gothra',
            'star_nakshatra', 'manglik_status', 'highest_education',
            'education_detail', 'occupation', 'employed_in', 'company',
            'annual_income', 'work_location', 'father_status', 'mother_status',
            'num_brothers', 'num_sisters', 'family_type', 'family_status',
            'family_location', 'about', 'hobbies',
        }
        for field in allowed_fields:
            if field in request.data:
                setattr(member, field, request.data[field])
        member.save()

        if any(f in request.data for f in profile_fields):
            profile, _ = MemberProfile.objects.get_or_create(member=member)
            for field in profile_fields:
                if field in request.data:
                    setattr(profile, field, request.data[field])
            profile.save()

        preference_fields = {
            'pref_age_min', 'pref_age_max', 'pref_height_min', 'pref_height_max',
            'pref_religion', 'pref_caste', 'pref_location', 'pref_education',
            'pref_occupation', 'pref_marital_status', 'pref_about',
        }
        preference_model_map = {
            'pref_age_min': 'preferred_age_min', 'pref_age_max': 'preferred_age_max',
            'pref_height_min': 'preferred_height_min', 'pref_height_max': 'preferred_height_max',
            'pref_religion': 'preferred_religion', 'pref_caste': 'preferred_caste',
            'pref_location': 'preferred_location', 'pref_education': 'preferred_education',
            'pref_occupation': 'preferred_occupation', 'pref_marital_status': 'preferred_marital_status',
            'pref_about': 'additional_expectations',
        }
        if any(f in request.data for f in preference_fields):
            preference, _ = MemberPreference.objects.get_or_create(member=member)
            for wire_name, model_name in preference_model_map.items():
                if wire_name in request.data:
                    setattr(preference, model_name, request.data[wire_name])
            preference.save()

        after = MemberSerializer(member, context={'request': request}).data
        audit(
            request, request.user, action='MEMBER_EDITED', module='members',
            target_type='MEMBER', target_id=member.pk, old_data=before, new_data=after,
        )
        return ApiResponse(data=after)

class AdminTicketListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = None

    def get(self, request):
        queryset = SupportTicket.objects.select_related('member', 'category')
        queryset = apply_scope_filter(request.user, queryset, branch_path='member__branch')
        for field in ('status', 'priority', 'source'):
            value = request.query_params.get(field)
            if value:
                queryset = queryset.filter(**{field: value})
        assigned_to = request.query_params.get('assigned_to')
        if assigned_to:
            queryset = queryset.filter(Q(claimed_by_admin_id=assigned_to) | Q(claimed_by_super_admin_id=assigned_to))
        return paginated_response(
            request, queryset, SupportTicketSerializer,
            context={'include_contact': True},
        )

def _ticket_actor_kwargs(actor):
    return {
        str(AccountType.ADMIN): {'changed_by_admin': actor},
        str(AccountType.SUPER_ADMIN): {'changed_by_super_admin': actor},
    }[str(actor.account_type)]

def _verification_actor_kwargs(actor):
    return {
        str(AccountType.ADMIN): {'changed_by_admin': actor},
        str(AccountType.SUPER_ADMIN): {'changed_by_super_admin': actor},
    }[str(actor.account_type)]

class AdminTicketDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = None
    parser_classes = (JSONParser, FormParser, MultiPartParser)

    def _ticket(self, ticket_id, lock=False):
        queryset = SupportTicket.objects.select_related('member', 'category')
        if lock:
            queryset = queryset.select_for_update(of=('self',))
        ticket = get_object_or_404(queryset, pk=ticket_id)
        check_object_scope(self.request.user, ticket, branch_path='member__branch')
        return ticket

    def get(self, request, ticket_id):
        ticket = self._ticket(ticket_id)
        data = SupportTicketSerializer(
            ticket,
            context={'request': request, 'include_contact': True, 'include_replies': True},
        ).data
        data['internal_notes'] = TicketInternalNoteSerializer(
            ticket.internal_notes.select_related('admin', 'super_admin'), many=True
        ).data
        return ApiResponse(data=data)

    @transaction.atomic
    def patch(self, request, ticket_id):
        ticket = self._ticket(ticket_id, lock=True)
        old = {'status': ticket.status, 'priority': ticket.priority, 'assignee': str(ticket.claimed_by_admin_id or ticket.claimed_by_super_admin_id or '')}
        if 'assigned_to' in request.data:
            if not request.user.has_admin_permission('tickets.assign'):
                raise PermissionDenied('Ticket assignment permission is required.')
            result = self._assign(request, ticket, request.data.get('assigned_to'))
            if result is not None:
                return result
        if 'priority' in request.data:
            if request.data['priority'] not in dict(SupportTicket.Priority.choices):
                return bad_request('Invalid ticket priority.')
            ticket.priority = request.data['priority']
        if 'status' in request.data and request.data['status'] != ticket.status:
            if request.data['status'] not in dict(SupportTicket.Status.choices):
                return bad_request('Invalid ticket status.')
            previous = ticket.status
            ticket.status = request.data['status']
            now = timezone.now()
            if ticket.status == SupportTicket.Status.RESOLVED:
                ticket.resolved_at = now
            if ticket.status == SupportTicket.Status.CLOSED:
                ticket.closed_at = now
            TicketStatusHistory.objects.create(
                ticket=ticket,
                old_status=previous,
                new_status=ticket.status,
                reason=str(request.data.get('reason', '')),
                **_ticket_actor_kwargs(request.user),
            )
        ticket.save()
        audit(
            request, request.user, action='TICKET_UPDATED', module='tickets',
            target_type='SUPPORT_TICKET', target_id=ticket.pk, old_data=old,
            new_data={'status': ticket.status, 'priority': ticket.priority, 'assignee': str(ticket.claimed_by_admin_id or ticket.claimed_by_super_admin_id or '')},
        )
        return ApiResponse(data=SupportTicketSerializer(ticket, context={'include_contact': True}).data)

    def _assign(self, request, ticket, assignee_id):
        if not assignee_id:
            TicketAssignment.objects.filter(ticket=ticket, is_current=True).update(
                is_current=False, unassigned_at=timezone.now()
            )
            ticket.claimed_by_admin = None
            ticket.claimed_by_super_admin = None
            if ticket.status == SupportTicket.Status.ASSIGNED:
                ticket.status = SupportTicket.Status.UNASSIGNED
            return None
        return bad_request('Customer support agents no longer supported.')
        return None

    @transaction.atomic
    def post(self, request, ticket_id):
        action = request.query_params.get('action') or request.data.get('action') or 'reply'
        ticket = self._ticket(ticket_id, lock=True)
        if action == 'note':
            return self._note(request, ticket)
        return self._reply(request, ticket)

    def _reply(self, request, ticket):
        if not request.user.has_admin_permission('tickets.reply'):
            raise PermissionDenied('Ticket reply permission is required.')
        serializer = TicketReplyInputSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        reply_values = {
            'ticket': ticket,
            'message': serializer.validated_data['message'],
            'is_public': True,
        }
        account_type_str = str(getattr(request.user, 'account_type', ''))
        if account_type_str == AccountType.SUPER_ADMIN:
            reply_values['super_admin_sender'] = request.user
        elif account_type_str == AccountType.ADMIN:
            reply_values['admin_sender'] = request.user
        else:
            reply_values['member_sender'] = request.user

        reply = SupportTicketReply.objects.create(**reply_values)
        attachment = serializer.validated_data.get('attachment')
        if attachment:
            if account_type_str == AccountType.SUPER_ADMIN:
                create_ticket_attachment(
                    ticket=ticket, reply=reply, upload=attachment, super_admin=request.user
                )
            elif account_type_str == AccountType.ADMIN:
                create_ticket_attachment(
                    ticket=ticket, reply=reply, upload=attachment, admin=request.user
                )
            else:
                create_ticket_attachment(
                    ticket=ticket, reply=reply, upload=attachment, member=request.user
                )
        now = timezone.now()
        if ticket.first_response_at is None:
            ticket.first_response_at = now
        ticket.last_reply_at = now
        old_status = ticket.status
        ticket.status = SupportTicket.Status.WAITING_FOR_MEMBER
        ticket.save(update_fields=('first_response_at', 'last_reply_at', 'status', 'updated_at'))
        if old_status != ticket.status:
            TicketStatusHistory.objects.create(
                ticket=ticket, old_status=old_status, new_status=ticket.status,
                reason='Admin replied', **_ticket_actor_kwargs(request.user),
            )
        if ticket.member_id:
            notify(
                ticket.member, notification_type='TICKET_REPLIED',
                title=f'New reply on {ticket.ticket_number}', message=ticket.subject,
                related_object=ticket, priority=ticket.priority,
            )
        audit(
            request, request.user, action='TICKET_REPLIED', module='tickets',
            target_type='SUPPORT_TICKET', target_id=ticket.pk,
        )
        return ApiResponse(
            data=SupportTicketReplySerializer(reply).data,
            status=status.HTTP_201_CREATED,
        )

    def _note(self, request, ticket):
        if not request.user.has_admin_permission('tickets.note'):
            raise PermissionDenied('Internal-note permission is required.')
        note = str(request.data.get('note', '')).strip()
        if not note:
            return bad_request('note is required.')
        values = {'ticket': ticket, 'note': note}
        account_type_str = str(getattr(request.user, 'account_type', ''))
        if account_type_str == AccountType.SUPER_ADMIN:
            values['super_admin'] = request.user
        elif account_type_str == AccountType.ADMIN:
            values['admin'] = request.user
        else:
            return bad_request('Unsupported account type for internal notes.')
        internal_note = TicketInternalNote.objects.create(**values)
        audit(
            request, request.user, action='TICKET_INTERNAL_NOTE_ADDED', module='tickets',
            target_type='SUPPORT_TICKET', target_id=ticket.pk,
        )
        return ApiResponse(
            data=TicketInternalNoteSerializer(internal_note).data,
            status=status.HTTP_201_CREATED,
        )

class AdminAssigneeListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'tickets.assign'

    def get(self, _request):
        return ApiResponse(data=[])



class AdminVerificationListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'verification.view_all'

    def get(self, request):
        requested_type = request.query_params.get('verification_type', '').upper()
        if requested_type:
            valid_types = {choice for choice, _label in ProfileVerificationRequest.VerificationType.choices}
            if requested_type not in valid_types:
                return bad_request('Unsupported verification_type filter.')

        if requested_type == ProfileVerificationRequest.VerificationType.FULL_PROFILE:
            self._sync_full_profile_verifications()

        queryset = ProfileVerificationRequest.objects.select_related('member')
        queryset = apply_scope_filter(request.user, queryset, branch_path='member__branch')
        requested_status = request.query_params.get('status')
        if requested_status:
            queryset = queryset.filter(status=requested_status)
        if requested_type:
            queryset = queryset.filter(verification_type=requested_type)
        return paginated_response(request, queryset, ProfileVerificationSerializer)

    @staticmethod
    def _sync_full_profile_verifications():
        from apps.accounts.models import Member
        active_statuses = {
            ProfileVerificationRequest.Status.PENDING_REVIEW,
            ProfileVerificationRequest.Status.IN_REVIEW,
            ProfileVerificationRequest.Status.CHANGES_REQUESTED,
        }
        pending_members = Member.objects.filter(
            profile_status=Member.VerificationStatus.PENDING_REVIEW,
            is_active=True,
            deleted_at__isnull=True,
        ).exclude(
            verification_requests__verification_type=ProfileVerificationRequest.VerificationType.FULL_PROFILE,
            verification_requests__status__in=active_statuses,
        ).distinct()
        ProfileVerificationRequest.objects.bulk_create([
            ProfileVerificationRequest(
                member=member,
                verification_type=ProfileVerificationRequest.VerificationType.FULL_PROFILE,
            )
            for member in pending_members
        ], ignore_conflicts=True)

class AdminVerificationDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'verification.view_all'

    def get(self, _request, verification_id):
        verification = get_object_or_404(ProfileVerificationRequest.objects.select_related('member'), pk=verification_id)
        check_object_scope(_request.user, verification, branch_path='member__branch')
        return ApiResponse(data=ProfileVerificationSerializer(verification).data)

    @transaction.atomic
    def post(self, request, verification_id):
        verification = get_object_or_404(
            ProfileVerificationRequest.objects.select_for_update().select_related('member'), pk=verification_id
        )
        check_object_scope(request.user, verification, branch_path='member__branch')
        action = request.data.get('action')
        if action == 'assign':
            if not request.user.has_admin_permission('verification.assign'):
                raise PermissionDenied('Verification assignment permission is required.')
            staff = get_object_or_404(Staff, pk=request.data.get('staff_id'), is_active=True, deleted_at__isnull=True)
            ProfileVerificationAssignment.objects.filter(
                verification_request=verification, is_current=True
            ).update(is_current=False, completed_at=timezone.now())
            assignment_values = {
                'verification_request': verification,
                'assigned_to_staff': staff,
            }
            if str(request.user.account_type) == AccountType.SUPER_ADMIN:
                assignment_values['assigned_by_super_admin'] = request.user
            else:
                assignment_values['assigned_by_admin'] = request.user
            ProfileVerificationAssignment.objects.create(**assignment_values)
            old_status = verification.status
            verification.status = ProfileVerificationRequest.Status.PENDING_REVIEW
            verification.save(update_fields=('status', 'updated_at'))
            ProfileVerificationHistory.objects.create(
                verification_request=verification, old_status=old_status,
                new_status=verification.status, reason='Assigned for staff review',
                **_verification_actor_kwargs(request.user),
            )
            notify(
                staff, notification_type='PROFILE_ASSIGNED', title='Profile verification assigned',
                message=verification.member.get_full_name(), related_object=verification,
                priority=verification.priority,
            )
            audit(
                request, request.user, action='PROFILE_VERIFICATION_ASSIGNED', module='verification',
                target_type='PROFILE_VERIFICATION', target_id=verification.pk,
                new_data={'staff_id': str(staff.pk)},
            )
            return ApiResponse(data=ProfileVerificationSerializer(verification).data)
        if action in {'approve', 'reject'}:
            permission = 'verification.approve' if action == 'approve' else 'verification.reject'
            if not request.user.has_admin_permission(permission):
                raise PermissionDenied(f'{permission} permission is required.')
            reviewable_statuses = {
                ProfileVerificationRequest.Status.PENDING_REVIEW,
                ProfileVerificationRequest.Status.IN_REVIEW,
                ProfileVerificationRequest.Status.CHANGES_REQUESTED,
            }
            if verification.status not in reviewable_statuses:
                return bad_request('This verification has already been completed.')
            return _review_verification(request, verification, action)
        return bad_request('Unsupported verification action.')

class StaffVerificationListView(ScopedAPIView):
    allowed_account_types = (AccountType.ADMIN,)
    required_permission = 'verification.view_assigned'

    def get(self, request):
        queryset = ProfileVerificationRequest.objects.filter(
            assignments__assigned_to_staff=request.user,
            assignments__is_current=True,
        ).select_related('member').distinct()
        queryset = apply_scope_filter(request.user, queryset, branch_path='member__branch')
        return paginated_response(request, queryset, ProfileVerificationSerializer)

class StaffVerificationDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.ADMIN,)
    required_permission = 'verification.view_assigned'

    def _verification(self, request, verification_id, lock=False):
        queryset = ProfileVerificationRequest.objects.select_related('member').filter(
            assignments__assigned_to_staff=request.user,
            assignments__is_current=True,
        )
        if lock:
            queryset = queryset.select_for_update()
        verification = get_object_or_404(queryset, pk=verification_id)
        check_object_scope(request.user, verification, branch_path='member__branch')
        return verification

    def get(self, request, verification_id):
        return ApiResponse(data=ProfileVerificationSerializer(self._verification(request, verification_id)).data)

    @transaction.atomic
    def post(self, request, verification_id):
        verification = self._verification(request, verification_id, lock=True)
        action = request.data.get('action')
        if action not in {'approve', 'reject', 'escalate', 'start_review'}:
            return bad_request('Unsupported verification action.')
        permission = {
            'approve': 'verification.approve',
            'reject': 'verification.reject',
            'escalate': 'verification.escalate',
            'start_review': 'verification.review',
        }[action]
        if not request.user.has_admin_permission(permission):
            raise PermissionDenied('You do not have permission for this review action.')
        return _review_verification(request, verification, action)

def activate_pending_membership_if_eligible(member):
    if getattr(member, 'are_verification_checks_passed', False):
        try:
            from datetime import timedelta
            membership = member.membership
            if getattr(membership, 'status', '') == 'PENDING_VERIFICATION':
                membership.status = 'ACTIVE'
                membership.start_date = timezone.now()
                duration_days = getattr(membership.plan, 'duration_days', 30)
                membership.end_date = membership.start_date + timedelta(days=duration_days)
                membership.is_active = True
                membership.save(update_fields=['status', 'start_date', 'end_date', 'is_active'])
                
                # Update user premium status
                member.is_premium = True
                member.save(update_fields=['is_premium', 'updated_at'])
        except Exception:
            pass

def _review_verification(request, verification, action):
    reason = str(request.data.get('reason', '')).strip()
    if action in {'reject', 'escalate'} and not reason:
        return bad_request(f'A reason is required to {action} a verification.')
    status_map = {
        'start_review': ProfileVerificationRequest.Status.IN_REVIEW,
        'approve': ProfileVerificationRequest.Status.APPROVED,
        'reject': ProfileVerificationRequest.Status.REJECTED,
        'escalate': ProfileVerificationRequest.Status.CHANGES_REQUESTED,
    }
    old_status = verification.status
    new_status = status_map[action]
    now = timezone.now()
    verification.status = new_status
    if action == 'approve':
        verification.reviewed_at = now
        verification.approved_at = now
        verification.rejected_at = None
        verification.rejection_reason = ''
        member = verification.member
        if verification.verification_type == ProfileVerificationRequest.VerificationType.FULL_PROFILE:
            member.profile_status = Member.ProfileStatus.APPROVED
        elif verification.verification_type == ProfileVerificationRequest.VerificationType.PROFILE_PHOTO:
            from apps.profiles.services.photo_management import review_all_pending_profile_photos

            review_all_pending_profile_photos(member=member, reviewer=request.user, approve=True)
            member.photo_status = Member.PhotoStatus.APPROVED
        elif verification.verification_type == ProfileVerificationRequest.VerificationType.IDENTITY_DOCUMENT:
            member.document_status = Member.DocumentStatus.APPROVED
            MemberDocument.objects.filter(
                member=member,
                status=MemberDocument.Status.PENDING,
            ).update(
                status=MemberDocument.Status.APPROVED,
                reviewed_at=now,
                reviewed_by_id=request.user.pk,
                rejection_reason='',
            )
        elif verification.verification_type == ProfileVerificationRequest.VerificationType.PHONE:
            member.is_mobile_verified = True
        elif verification.verification_type == ProfileVerificationRequest.VerificationType.EMAIL:
            member.is_email_verified = True
        member.save()
        
        # Trigger pending membership activation check
        activate_pending_membership_if_eligible(member)

    elif action == 'reject':
        verification.reviewed_at = now
        verification.rejected_at = now
        verification.rejection_reason = reason
        member = verification.member
        if verification.verification_type == ProfileVerificationRequest.VerificationType.FULL_PROFILE:
            member.profile_status = Member.ProfileStatus.REJECTED
            member.save(update_fields=('profile_status', 'updated_at'))
        elif verification.verification_type == ProfileVerificationRequest.VerificationType.PROFILE_PHOTO:
            from apps.profiles.services.photo_management import review_all_pending_profile_photos

            review_all_pending_profile_photos(
                member=member,
                reviewer=request.user,
                approve=False,
                reason=reason,
            )
            member.photo_status = Member.PhotoStatus.REJECTED
            member.save(update_fields=('photo_status', 'updated_at'))
        elif verification.verification_type == ProfileVerificationRequest.VerificationType.IDENTITY_DOCUMENT:
            member.document_status = Member.DocumentStatus.REJECTED
            MemberDocument.objects.filter(
                member=member,
                status=MemberDocument.Status.PENDING,
            ).update(
                status=MemberDocument.Status.REJECTED,
                rejection_reason=reason,
                reviewed_at=now,
                reviewed_by_id=request.user.pk,
            )
            member.save(update_fields=('document_status', 'updated_at'))
    elif action == 'escalate':
        verification.escalation_reason = reason
    verification.save()
    ProfileVerificationHistory.objects.create(
        verification_request=verification,
        old_status=old_status,
        new_status=new_status,
        reason=reason,
        **_verification_actor_kwargs(request.user),
    )
    if action != 'start_review':
        ProfileVerificationAssignment.objects.filter(
            verification_request=verification,
            is_current=True,
        ).update(is_current=False, completed_at=now)
    notify(
        verification.member,
        notification_type={
            'approve': 'PROFILE_APPROVED',
            'reject': 'PROFILE_REJECTED',
            'escalate': 'PROFILE_ESCALATED',
            'start_review': 'PROFILE_IN_REVIEW',
        }[action],
        title=f'Profile verification {new_status.lower().replace("_", " ")}',
        message=reason or 'Your verification status has been updated.',
        related_object=verification,
        priority=verification.priority,
    )
    audit(
        request, request.user, action=f'PROFILE_{new_status.upper()}', module='verification',
        target_type='PROFILE_VERIFICATION', target_id=verification.pk,
        old_data={'status': old_status}, new_data={'status': new_status, 'reason': reason},
    )
    return ApiResponse(data=ProfileVerificationSerializer(verification).data)

class StaffDashboardView(ScopedAPIView):
    allowed_account_types = (AccountType.ADMIN,)

    def get(self, request):
        from apps.core.models import WorkAssignment, Notification
        from apps.accounts.models import StaffActivityLog
        from django.utils import timezone
        from datetime import timedelta
        
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=7)
        
        assigned = WorkAssignment.objects.filter(assigned_to_staff=request.user, status='ASSIGNED').count()
        in_progress = WorkAssignment.objects.filter(assigned_to_staff=request.user, status='IN_PROGRESS').count()
        due_today = WorkAssignment.objects.filter(assigned_to_staff=request.user, status__in=('ASSIGNED', 'IN_PROGRESS'), due_at__date=now.date()).count()
        overdue = WorkAssignment.objects.filter(assigned_to_staff=request.user, status__in=('ASSIGNED', 'IN_PROGRESS'), due_at__lt=now).count()
        escalated = WorkAssignment.objects.filter(assigned_to_staff=request.user, status='ESCALATED').count()
        completed_today = WorkAssignment.objects.filter(assigned_to_staff=request.user, status='COMPLETED', completed_at__gte=today_start).count()
        completed_week = WorkAssignment.objects.filter(assigned_to_staff=request.user, status='COMPLETED', completed_at__gte=week_start).count()

        recent_assignments = WorkAssignment.objects.filter(assigned_to_staff=request.user).order_by('-created_at')[:5]
        recent_activity = StaffActivityLog.objects.filter(actor_id=request.user.id).order_by('-created_at')[:5]
        unread_notifications = Notification.objects.filter(staff_recipient=request.user, is_read=False).count()

        recent_assigns_serialized = WorkAssignmentSerializer(recent_assignments, many=True).data
        recent_acts_serialized = [{
            'id': str(log.id),
            'action': log.action,
            'description': log.description,
            'created_at': log.created_at,
        } for log in recent_activity]

        return ApiResponse(data={
            'summary': {
                'assigned': assigned,
                'in_progress': in_progress,
                'due_today': due_today,
                'overdue': overdue,
                'escalated': escalated,
                'completed_today': completed_today,
                'completed_week': completed_week,
            },
            'recent_assignments': recent_assigns_serialized,
            'recent_activity': recent_acts_serialized,
            'unread_notifications': unread_notifications,
        })

class AdministrativeNotificationListView(ScopedAPIView):
    allowed_account_types = (
        AccountType.SUPER_ADMIN,
        AccountType.ADMIN,
        AccountType.ADMIN,
    )

    def get(self, request):
        recipient_field = {
            AccountType.SUPER_ADMIN: 'super_admin_recipient',
            AccountType.ADMIN: 'admin_recipient',
            AccountType.ADMIN: 'staff_recipient',
        }[str(request.user.account_type)]
        queryset = Notification.objects.filter(**{recipient_field: request.user})
        from .serializers import NotificationSerializer
        return paginated_response(request, queryset, NotificationSerializer)

class AdminActivityListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    def get(self, request):
        from django.db.models import Q

        if str(request.user.account_type) == AccountType.SUPER_ADMIN:
            model_pairs = (
                ('SUPER_ADMIN', SuperAdminActivityLog),
                ('ADMIN', AdminActivityLog),
                ('ADMIN', StaffActivityLog),
            )
        else:
            if not (
                request.user.has_admin_permission('staff.activity')
                or request.user.has_admin_permission('support_agents.activity')
            ):
                raise PermissionDenied()
            model_pairs = (('ADMIN', StaffActivityLog),)

        role_filter = request.query_params.get('role', '').upper()
        search = request.query_params.get('search', '').strip()
        module_filter = request.query_params.get('module', '').strip().lower()

        try:
            page_size = max(1, min(int(request.query_params.get('page_size', 20)), 20))
        except (TypeError, ValueError):
            page_size = 20
        try:
            page_number = max(1, int(request.query_params.get('page', 1)))
        except (TypeError, ValueError):
            page_number = 1

        rows = []
        for role, model in model_pairs:
            if role_filter and role != role_filter:
                continue
            qs = model.objects.all()
            if str(request.user.account_type) != AccountType.SUPER_ADMIN:
                qs = qs.filter(actor_id=request.user.pk)
            search_filters = Q()
            if search:
                search_filters = (
                    Q(action__icontains=search) |
                    Q(module__icontains=search) |
                    Q(description__icontains=search)
                )
                if hasattr(model, 'actor_name'):
                    search_filters |= Q(actor_name__icontains=search)
                qs = qs.filter(search_filters)
            if module_filter:
                if module_filter in {'photo', 'photos'}:
                    qs = qs.filter(Q(module__iexact='photos') | Q(module__iexact='members', action__icontains='PHOTO'))
                elif module_filter in {'document', 'documents'}:
                    qs = qs.filter(Q(module__iexact='documents') | Q(module__iexact='verification', action__icontains='DOCUMENT'))
                elif module_filter in {'member', 'members'}:
                    qs = qs.filter(module__iexact='members')
                else:
                    qs = qs.filter(module__iexact=module_filter)
            for item in qs.order_by('-created_at')[:200]:
                rows.append((role, item))

        rows.sort(key=lambda pair: pair[1].created_at, reverse=True)
        total = len(rows)
        start = (page_number - 1) * page_size
        end = start + page_size
        page_rows = rows[start:end]

        data = []
        for role, item in page_rows:
            target_account_info = None
            try:
                from apps.accounts.models import Member, MemberDocument
                from apps.profiles.models import ProfilePhoto
                from apps.core.models import SupportTicket

                t_type = str(getattr(item, 'target_type', '') or '').upper()
                t_id = str(getattr(item, 'target_id', '') or '')
                desc = getattr(item, 'description', '') or ''

                if t_type in {'MEMBER', 'USER', 'MEMBERPROFILE'} and t_id:
                    m = Member.objects.filter(pk=t_id).first()
                    if m:
                        target_account_info = {'full_name': m.get_full_name(), 'email': m.email, 'type': 'Member Profile'}
                elif t_type in {'ADMIN', 'STAFF', 'SUPER_ADMIN', 'ACCOUNT'} and t_id:
                    from apps.accounts.models import Admin, Staff, SuperAdmin
                    adm = Admin.objects.filter(pk=t_id).first() or Staff.objects.filter(pk=t_id).first() or SuperAdmin.objects.filter(pk=t_id).first()
                    if adm:
                        target_account_info = {'full_name': adm.get_full_name(), 'email': adm.email, 'type': f'{t_type.capitalize()} Account'}
                elif 'TICKET' in t_type or 'ticket' in desc.lower():
                    st = SupportTicket.objects.filter(pk=t_id).first()
                    if not st and desc:
                        for tok in desc.split():
                            if tok.startswith('MAT-') or tok.startswith('TCK-'):
                                st = SupportTicket.objects.filter(ticket_number=tok.strip('(),.')).first()
                                if st: break
                    if st and st.member:
                        target_account_info = {'full_name': st.member.get_full_name(), 'email': st.member.email, 'ticket_number': st.ticket_number, 'type': 'Support Ticket'}
                elif 'DOCUMENT' in t_type or 'doc' in desc.lower():
                    doc = MemberDocument.objects.filter(pk=t_id).first()
                    if doc and doc.member:
                        target_account_info = {'full_name': doc.member.get_full_name(), 'email': doc.member.email, 'type': 'Document Verification'}
                elif 'PHOTO' in t_type or 'photo' in desc.lower():
                    photo = ProfilePhoto.objects.filter(pk=t_id).first()
                    if photo and photo.user:
                        target_account_info = {'full_name': photo.user.get_full_name(), 'email': photo.user.email, 'type': 'Photo Verification'}
            except Exception:
                pass

            readable_desc = item.description
            if not readable_desc:
                action_str = str(item.action or '').upper()
                if 'APPROVE_PROFILE' in action_str:
                    readable_desc = 'Approved member profile details'
                elif 'APPROVE_PHOTO' in action_str:
                    readable_desc = 'Approved member profile photo'
                elif 'REJECT_PHOTO' in action_str:
                    readable_desc = 'Rejected member profile photo'
                elif 'VERIFY_DOCUMENT' in action_str or 'DOCUMENT_APPROVED' in action_str:
                    readable_desc = 'Verified member document'
                elif 'ACTIVATE' in action_str:
                    readable_desc = 'Activated member account'
                elif 'DEACTIVATE' in action_str or 'SUSPEND' in action_str:
                    readable_desc = 'Suspended member account'
                else:
                    readable_desc = action_str.replace('_', ' ').title()

            data.append({
                'id': str(item.pk),
                'actor_id': str(item.actor_id),
                'actor_name': item.actor_name or '',
                'actor_role': item.actor_role or role,
                'role': role,
                'action': item.action,
                'module': item.module,
                'record_id': item.target_id,
                'description': readable_desc,
                'target_account': target_account_info,
                'ip_address': item.ip_address,
                'user_agent': item.user_agent,
                'latitude': item.latitude,
                'longitude': item.longitude,
                'city': item.city or '',
                'country': item.country or '',
                'was_successful': True,
                'created_at': item.created_at,
            })
        return ApiResponse(data={
            'count': total,
            'page': page_number,
            'page_size': page_size,
            'num_pages': max(1, (total + page_size - 1) // page_size),
            'results': data,
        })

class _PaymentWireSerializer(serializers.Serializer):
    def to_representation(self, payment):
        member = payment.member
        return {
            'id': str(payment.pk),
            'user': member.get_full_name() if member else 'Deleted member',
            'email': member.email if member else '',
            'plan': payment.plan.name if payment.plan_id else 'No plan',
            'amount': str(payment.amount),
            'status': payment.status,
            'date': payment.created_at,
            'gateway': payment.gateway,
            'reference': payment.gateway_reference or str(payment.client_reference),
            'refund_status': payment.refund_status,
            'refunded_amount': str(payment.refunded_amount),
        }

class AdminPaymentListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN, AccountType.ADMIN)
    required_permission = 'payments.view'

    def get(self, request):
        queryset = PaymentOrder.objects.select_related('user', 'membership_plan').order_by('-created_at')
        
        status_filter = request.query_params.get('status', '').strip()
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(user__email__icontains=search)
                | Q(user__first_name__icontains=search)
                | Q(user__last_name__icontains=search)
                | Q(razorpay_order_id__icontains=search)
                | Q(internal_order_number__icontains=search)
            )
            
        is_limited = request.user.account_type == AccountType.ADMIN
        
        class PaymentOrderWireSerializer(serializers.Serializer):
            def to_representation(self, obj):
                member = obj.user
                rep = {
                    'id': str(obj.pk),
                    'user': member.get_full_name() if member else 'Deleted user',
                    'email': member.email if member else '',
                    'plan': obj.membership_plan.name if obj.membership_plan else 'No plan',
                    'amount': str(obj.amount),
                    'currency': obj.currency,
                    'status': obj.status,
                    'internal_order_number': obj.internal_order_number,
                    'razorpay_order_id': obj.razorpay_order_id,
                    'receipt': obj.receipt,
                    'date': obj.created_at.isoformat(),
                    'created_at': obj.created_at.isoformat(),
                    'updated_at': obj.updated_at.isoformat(),
                    'gateway': 'Razorpay' if obj.razorpay_order_id else 'Manual',
                    'reference': obj.razorpay_order_id or obj.internal_order_number,
                }
                if not is_limited:
                    rep['notes'] = getattr(obj, 'notes', None)
                return rep

        return paginated_response(request, queryset, PaymentOrderWireSerializer)

class AdminPaymentDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN, AccountType.ADMIN)
    required_permission = 'payments.view'

    def get(self, request, id):
        order = get_object_or_404(PaymentOrder.objects.select_related('user', 'membership_plan'), pk=id)
        txs = PaymentTransaction.objects.filter(payment_order=order)
        refunds = RefundRequest.objects.filter(payment_transaction__payment_order=order).prefetch_related('transactions')
        
        is_limited = request.user.account_type == AccountType.ADMIN
        
        txs_serialized = []
        for tx in txs:
            t_data = {
                'id': str(tx.pk),
                'razorpay_payment_id': tx.razorpay_payment_id,
                'amount': str(tx.amount),
                'currency': tx.currency,
                'method': tx.method,
                'status': tx.status,
                'bank': tx.bank,
                'wallet': tx.wallet,
                'card_network': tx.card_network,
                'card_last4': tx.card_last4,
                'created_at': tx.created_at.isoformat()
            }
            if not is_limited:
                t_data['safe_metadata'] = tx.safe_metadata
            txs_serialized.append(t_data)
            
        refunds_serialized = []
        for r in refunds:
            tx_ref = r.transactions.first()
            refunds_serialized.append({
                'id': str(r.pk),
                'razorpay_refund_id': tx_ref.razorpay_refund_id if tx_ref else None,
                'internal_refund_number': tx_ref.internal_refund_number if tx_ref else f'REQ-{r.pk.hex[:8].upper()}',
                'amount': str(r.requested_amount),
                'currency': order.currency,
                'status': r.status,
                'reason': r.reason,
                'customer_note': r.details,
                'admin_note': '' if is_limited else r.admin_note,
                'created_at': r.requested_at.isoformat()
            })

        data = {
            'id': str(order.pk),
            'internal_order_number': order.internal_order_number,
            'razorpay_order_id': order.razorpay_order_id,
            'amount': str(order.amount),
            'currency': order.currency,
            'status': order.status,
            'receipt': order.receipt,
            'created_at': order.created_at.isoformat(),
            'user': {
                'id': str(order.user.pk) if order.user else None,
                'full_name': order.user.get_full_name() if order.user else 'Deleted user',
                'email': order.user.email if order.user else '',
                'phone': order.user.mobile_number if order.user else ''
            },
            'plan': {
                'id': str(order.membership_plan.pk) if order.membership_plan else None,
                'name': order.membership_plan.name if order.membership_plan else 'No plan',
                'price': str(order.membership_plan.price) if order.membership_plan else '0.00',
                'duration_days': order.membership_plan.duration_days if order.membership_plan else 0
            },
            'transactions': txs_serialized,
            'refunds': refunds_serialized
        }
        if not is_limited:
            data['notes'] = {}
            
        return ApiResponse(data=data)

class AdminPaymentRefundView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'payments.refund'

    def post(self, request, id):
        from apps.core.services.razorpay_memberships import RazorpayMembershipService, RazorpayGatewayError
        from decimal import Decimal, ROUND_HALF_UP
        import uuid

        # 1. Fetch transaction and validate status (No DB lock open during network call)
        # The admin list exposes PaymentOrder ids, so resolve the captured transaction
        # from the order when a transaction id is not supplied directly.
        tx = PaymentTransaction.objects.filter(pk=id).first()
        if tx is None:
            order = get_object_or_404(PaymentOrder, pk=id)
            tx = PaymentTransaction.objects.filter(
                payment_order=order, status='captured'
            ).order_by('-created_at').first()
            if tx is None:
                return bad_request('This payment has no captured transaction to refund.')
        else:
            order = tx.payment_order

        if tx.status != 'captured':
            return bad_request('Only captured transactions can be refunded.')
            
        try:
            requested_amount = serializers.DecimalField(max_digits=12, decimal_places=2).run_validation(
                request.data.get('amount')
            )
        except serializers.ValidationError as exc:
            return bad_request('Enter a valid refund amount.', errors={'amount': exc.detail})
            
        if requested_amount <= 0:
            return bad_request('Refund amount must be greater than zero.')

        # 2. Acquire a short DB lock to calculate the remaining balance
        with transaction.atomic():
            tx_locked = PaymentTransaction.objects.select_for_update().get(pk=tx.pk)
            total_refunded = sum(rt.amount for rt in RefundTransaction.objects.filter(payment_transaction=tx_locked, status__in=('processed', 'processing')))
            available = tx_locked.amount - total_refunded

        if requested_amount > available:
            return bad_request(f'Refund amount exceeds remaining refundable balance of {available}.')

        reason = str(request.data.get('reason', '')).strip() or 'approved_customer_request'
        customer_note = str(request.data.get('customer_note', '')).strip()
        admin_note = str(request.data.get('admin_note', '')).strip()
        speed = str(request.data.get('speed', 'normal')).strip()

        # 3. Call Razorpay API OUTSIDE Django transaction
        if settings.RAZORPAY_DEMO_MODE:
            razorpay_refund_id = f'demo_ref_{uuid.uuid4().hex[:12]}'
            status_val = 'processed'
            gateway_res = {'id': razorpay_refund_id, 'status': 'processed'}
        else:
            try:
                amount_in_cents = int((requested_amount * 100).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
                gateway_res = RazorpayMembershipService.initiate_razorpay_refund(tx.razorpay_payment_id, amount_in_cents, speed)
                razorpay_refund_id = gateway_res['id']
                status_val = 'processed' if gateway_res.get('status') == 'processed' else 'processing'
            except Exception as exc:
                return bad_request(f'Failed to initiate refund on Razorpay: {str(exc)}')

        # 4. Save results inside another short transaction
        with transaction.atomic():
            tx_locked = PaymentTransaction.objects.select_for_update().get(pk=tx.pk)
            
            # Link to an existing request or create a new approved one
            rr = RefundRequest.objects.filter(payment_transaction=tx_locked, status='approved').first()
            if not rr:
                rr = RefundRequest.objects.create(
                    payment_transaction=tx_locked,
                    user=tx_locked.user,
                    requested_amount=requested_amount,
                    reason=reason,
                    details=customer_note,
                    status='approved',
                    reviewed_at=timezone.now(),
                    reviewed_by=request.user,
                    admin_note=admin_note
                )
            else:
                rr.reviewed_at = timezone.now()
                rr.reviewed_by = request.user
                rr.admin_note = admin_note
                rr.save()

            internal_ref_num = f'REF-{timezone.now().strftime("%Y%m%d")}-{uuid.uuid4().hex[:8].upper()}'
            
            # Create RefundTransaction
            rt = RefundTransaction.objects.create(
                refund_request=rr,
                payment_transaction=tx_locked,
                razorpay_refund_id=razorpay_refund_id,
                internal_refund_number=internal_ref_num,
                amount=requested_amount,
                currency=tx_locked.currency,
                status=status_val,
                processed_at=timezone.now() if status_val == 'processed' else None,
                safe_metadata=gateway_res
            )

            if status_val == 'processed':
                rr.status = 'processed'
                rr.save()

                total_refunded_updated = sum(r.amount for r in RefundTransaction.objects.filter(payment_transaction=tx_locked, status='processed'))
                if total_refunded_updated >= tx_locked.amount:
                    order.status = 'refunded'
                    order.save(update_fields=('status', 'updated_at'))
                    MembershipPurchase.objects.filter(payment_transaction=tx_locked, status='active').update(status='refunded')
                    MemberMembership.objects.filter(razorpay_payment_id=tx_locked.razorpay_payment_id).update(is_active=False, status=MemberMembership.MembershipStatus.EXPIRED)
                    Member.objects.filter(pk=tx_locked.user_id).update(is_premium=False)
                else:
                    order.status = 'partially_refunded'
                    order.save(update_fields=('status', 'updated_at'))
                    MembershipPurchase.objects.filter(payment_transaction=tx_locked, status='active').update(status='partially_refunded')
            else:
                rr.status = 'processing'
                rr.save()

        audit(
            request, request.user, action='REFUND_ISSUED', module='payments',
            target_type='PAYMENT_TRANSACTION', target_id=tx.pk,
            new_data={'amount': str(requested_amount), 'refund_id': str(rr.pk)},
        )
        
        return ApiResponse(
            data={'id': str(rr.pk), 'status': rr.status, 'amount': str(rr.requested_amount)},
            message='Refund initiated successfully.',
            status=status.HTTP_201_CREATED,
        )

class AdminRefundRequestApproveRejectView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'payments.refund'

    @transaction.atomic
    def post(self, request, id, action):
        refund = get_object_or_404(RefundRequest.objects.select_for_update(), pk=id)
        if refund.status != 'requested':
            return bad_request(f'Refund request is in status {refund.status} and cannot be processed.')
            
        admin_note = str(request.data.get('admin_note', '')).strip()
        refund.reviewed_by = request.user
        refund.reviewed_at = timezone.now()
        refund.admin_note = admin_note
        
        if action == 'reject':
            refund.status = 'rejected'
            refund.save()
            audit(
                request, request.user, action='REFUND_REQUEST_REJECTED', module='payments',
                target_type='REFUND_REQUEST', target_id=refund.pk,
                new_data={'admin_note': admin_note}
            )
            return ApiResponse(message='Refund request rejected.')
            
        # Approval step ONLY. Execution is a separate step (triggered via Issue Refund)
        refund.status = 'approved'
        refund.save()

        audit(
            request, request.user, action='REFUND_REQUEST_APPROVED', module='payments',
            target_type='REFUND_REQUEST', target_id=refund.pk,
            new_data={'amount': str(refund.requested_amount)}
        )
        
        return ApiResponse(
            data={'id': str(refund.pk), 'status': refund.status, 'amount': str(refund.requested_amount)},
            message='Refund request approved. Click Issue Refund to execute.',
            status=status.HTTP_200_OK,
        )

class AdminPaymentReconcileView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'payments.reconcile'

    def post(self, request, id):
        from apps.core.services.razorpay_memberships import RazorpayMembershipService
        import base64
        import json
        from urllib.request import Request, urlopen
        
        order = get_object_or_404(PaymentOrder, pk=id)
        fixes = []
        
        if settings.RAZORPAY_DEMO_MODE:
            if order.status == 'created':
                fixes.append("Order was in pending created state. No live Razorpay transaction found. Kept created.")
            return ApiResponse(data={'status': 'completed', 'fixes': fixes})
            
        try:
            if order.razorpay_order_id:
                url = f'https://api.razorpay.com/v1/orders/{order.razorpay_order_id}/payments'
                credentials = base64.b64encode(
                    f'{settings.RAZORPAY_KEY_ID}:{settings.RAZORPAY_KEY_SECRET}'.encode('utf-8')
                ).decode('ascii')
                req = Request(
                    url,
                    headers={'Authorization': f'Basic {credentials}', 'Content-Type': 'application/json'},
                    method='GET',
                )
                with urlopen(req, timeout=15) as res:
                    payment_list = json.loads(res.read().decode('utf-8'))
                    items = payment_list.get('items', [])
                    
                captured_pay = None
                for pay in items:
                    if pay.get('status') == 'captured':
                        captured_pay = pay
                        break
                        
                if captured_pay:
                    payment_id = captured_pay['id']
                    tx_exists = PaymentTransaction.objects.filter(razorpay_payment_id=payment_id).exists()
                    if not tx_exists or order.status != 'paid':
                        RazorpayMembershipService.activate_order_transactional(
                            order_id=order.razorpay_order_id,
                            payment_id=payment_id,
                            raw_payload=captured_pay
                        )
                        fixes.append(f"Detected captured payment {payment_id} on Razorpay but local order was unpaid. Activated membership.")
                else:
                    if order.status == 'created' and order.expires_at and order.expires_at <= timezone.now():
                        order.status = 'expired'
                        order.save(update_fields=('status', 'updated_at'))
                        fixes.append("Order expired.")
        except Exception as exc:
            return ApiResponse(success=False, message=f"Reconciliation failed: {str(exc)}", status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return ApiResponse(data={'status': 'completed', 'fixes': fixes})

class _MembershipWireSerializer(serializers.Serializer):
    def to_representation(self, membership):
        # Check the member's current active plan (prefetched to avoid N+1)
        prefetched = getattr(membership.member, '_active_memberships', None)
        if prefetched is not None:
            current = prefetched[0] if prefetched else None
        else:
            current = membership.member.memberships.filter(is_active=True).select_related('plan').first()
        return {
            'id': str(membership.pk),
            'user': member_summary(membership.member, include_contact=True),
            'plan': {
                'id': str(membership.plan_id),
                'name': membership.plan.name,
                'price': str(membership.plan.price),
            } if membership.plan_id else None,
            # current_plan reflects what the member has RIGHT NOW (may differ from this record's plan)
            'current_plan': {
                'name': current.plan.name if (current and current.plan) else 'Free',
                'is_active': current.is_active if current else False,
                'expires_at': current.end_date if current else None,
            },
            'start_date': membership.start_date,
            'end_date': membership.end_date,
            'is_active': membership.is_active,
        }

class AdminMemberSearchView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'members.manage'

    def get(self, request):
        from apps.accounts.models import Member
        from django.db.models import Q

        q = request.query_params.get('q', '').strip()
        queryset = Member.objects.filter(deleted_at__isnull=True)
        if q:
            queryset = queryset.filter(
                Q(email__icontains=q)
                | Q(first_name__icontains=q)
                | Q(last_name__icontains=q)
            )
        queryset = queryset.order_by('first_name')[:10]
        results = [
            {
                'id': str(m.pk),
                'full_name': m.get_full_name() or m.email,
                'email': m.email,
                'is_premium': m.is_premium,
                'account_status': m.account_status,
            }
            for m in queryset
        ]
        return ApiResponse(data={'results': results})

class AdminMembershipListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'members.view'

    def get(self, request):
        from django.db.models import Prefetch
        # Prefetch each member's current active plan to avoid N+1 queries
        active_m_qs = MemberMembership.objects.filter(is_active=True).select_related('plan')
        queryset = MemberMembership.objects.select_related('member', 'plan').prefetch_related(
            Prefetch('member__memberships', queryset=active_m_qs, to_attr='_active_memberships')
        )
        requested_status = request.query_params.get('status')
        if requested_status == 'active':
            queryset = queryset.filter(is_active=True)
        elif requested_status == 'expired':
            queryset = queryset.filter(is_active=False)
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(member__email__icontains=search)
                | Q(member__first_name__icontains=search)
                | Q(member__last_name__icontains=search)
            )
        return paginated_response(request, queryset.order_by('-start_date'), _MembershipWireSerializer)

class _MembershipRequestWireSerializer(serializers.Serializer):
    def to_representation(self, req):
        return {
            'id': str(req.pk),
            'user': member_summary(req.user, include_contact=True),
            'selected_plan': {
                'id': str(req.selected_plan_id),
                'name': req.selected_plan.name,
                'price': str(req.selected_plan.price),
                'slug': req.selected_plan.slug,
            },
            'status': req.status,
            'requested_at': req.requested_at,
            'approved_at': req.approved_at,
            'approved_by': administrative_summary(req.approved_by) if req.approved_by else None,
            'rejected_at': req.rejected_at,
            'rejected_by': administrative_summary(req.rejected_by) if req.rejected_by else None,
            'rejection_reason': req.rejection_reason,
            'start_date': req.start_date,
            'expiry_date': req.expiry_date,
            'is_active': req.is_active,
        }

class AdminMembershipRequestListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'members.view'

    def get(self, request):
        from apps.core.models import MembershipRequest
        queryset = MembershipRequest.objects.select_related('user', 'selected_plan')
        requested_status = request.query_params.get('status')
        if requested_status:
            queryset = queryset.filter(status=requested_status)
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(user__email__icontains=search)
                | Q(user__first_name__icontains=search)
                | Q(user__last_name__icontains=search)
            )
        return paginated_response(request, queryset.order_by('-requested_at'), _MembershipRequestWireSerializer)

class AdminMembershipRequestDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'members.manage'

    @transaction.atomic
    def patch(self, request, pk):
        from datetime import timedelta
        from apps.core.models import MembershipRequest
        req = get_object_or_404(MembershipRequest.objects.select_for_update(), pk=pk)
        
        action = request.data.get('action')
        if action not in ('approve', 'reject'):
            return ApiResponse(success=False, message="Action must be 'approve' or 'reject'.", status=status.HTTP_400_BAD_REQUEST)
        
        if req.status != 'pending':
            return ApiResponse(success=False, message="Request is not in pending status.", status=status.HTTP_400_BAD_REQUEST)
        
        actor = request.user
        now = timezone.now()
        
        if action == 'approve':
            plan = req.selected_plan
            # Deactivate any previous active memberships for this member
            MemberMembership.objects.filter(member=req.user, is_active=True).update(is_active=False)
            
            # Setup start and end dates
            start_date = now
            duration_days = getattr(plan, 'duration_days', 30)
            end_date = start_date + timedelta(days=duration_days) if duration_days else None
            
            # Update/create MemberMembership record
            membership, created = MemberMembership.objects.update_or_create(
                member=req.user,
                defaults={
                    'plan': plan,
                    'start_date': start_date,
                    'end_date': end_date,
                    'is_active': True
                }
            )
            
            # Activate premium flag on the user
            req.user.is_premium = True
            req.user.save(update_fields=['is_premium', 'updated_at'])
            
            # Update request status
            req.status = 'approved'
            req.approved_at = now
            req.approved_by_id = actor.pk
            req.start_date = start_date
            req.expiry_date = end_date
            req.is_active = True
            req.save(update_fields=['status', 'approved_at', 'approved_by_id', 'start_date', 'expiry_date', 'is_active'])
            
            audit(
                request, actor, action='MEMBERSHIP_APPROVED', module='memberships',
                target_type='MEMBER', target_id=req.user.pk,
                new_data={'plan_slug': plan.slug, 'duration_days': duration_days, 'request_id': str(req.pk)}
            )
            
            return ApiResponse(message="Membership request approved and plan activated successfully.")
            
        elif action == 'reject':
            rejection_reason = request.data.get('rejection_reason', '').strip()
            if not rejection_reason:
                return ApiResponse(success=False, message="Rejection reason is required.", status=status.HTTP_400_BAD_REQUEST)
                
            req.status = 'rejected'
            req.rejected_at = now
            req.rejected_by_id = actor.pk
            req.rejection_reason = rejection_reason
            req.is_active = False
            req.save(update_fields=['status', 'rejected_at', 'rejected_by_id', 'rejection_reason', 'is_active'])
            
            audit(
                request, actor, action='MEMBERSHIP_REJECTED', module='memberships',
                target_type='MEMBER', target_id=req.user.pk,
                new_data={'reason': rejection_reason, 'request_id': str(req.pk)}
            )
            
            return ApiResponse(message="Membership request rejected.")

class AdminDirectMembershipView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'members.manage'

    @transaction.atomic
    def post(self, request):
        from datetime import timedelta
        from apps.core.models import MembershipRequest
        user_id = request.data.get('user_id')
        plan_slug = request.data.get('plan_slug')
        action = request.data.get('action') # activate, extend, cancel
        duration_days = request.data.get('duration_days')

        if not user_id or not action:
            return ApiResponse(
                success=False,
                message=f"user_id and action are required. (received user_id={user_id!r}, action={action!r}, plan_slug={plan_slug!r})",
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            target_user = Member.objects.select_for_update().get(pk=user_id)
        except (Member.DoesNotExist, ValueError):
            return ApiResponse(
                success=False,
                message=f"No member found for user_id '{user_id}'.",
                status=status.HTTP_400_BAD_REQUEST,
            )
        actor = request.user
        now = timezone.now()

        if action == 'activate':
            if not plan_slug:
                return ApiResponse(success=False, message="plan_slug is required for activation.", status=status.HTTP_400_BAD_REQUEST)
            try:
                plan = MembershipPlan.objects.get(slug=plan_slug)
            except (MembershipPlan.DoesNotExist, ValueError):
                return ApiResponse(
                    success=False,
                    message=f"No membership plan found for plan_slug '{plan_slug}'.",
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Deactivate previous active memberships
            MemberMembership.objects.filter(member=target_user, is_active=True).update(is_active=False)
            
            start_date = now
            days = int(duration_days) if duration_days else getattr(plan, 'duration_days', 30)
            end_date = start_date + timedelta(days=days) if days else None
            
            # Update/create MemberMembership record
            MemberMembership.objects.update_or_create(
                member=target_user,
                defaults={
                    'plan': plan,
                    'start_date': start_date,
                    'end_date': end_date,
                    'is_active': True
                }
            )
            
            target_user.is_premium = True
            target_user.save(update_fields=['is_premium', 'updated_at'])
            
            # Create approved MembershipRequest for history audit trail
            MembershipRequest.objects.create(
                user=target_user,
                selected_plan=plan,
                status='approved',
                requested_at=now,
                approved_at=now,
                approved_by_id=actor.pk,
                start_date=start_date,
                expiry_date=end_date,
                is_active=True
            )
            
            audit(
                request, actor, action='MEMBERSHIP_DIRECT_ACTIVATED', module='memberships',
                target_type='MEMBER', target_id=target_user.pk,
                new_data={'plan_slug': plan.slug, 'duration_days': days}
            )

            create_notification(
                recipient=target_user,
                type='MEMBERSHIP_ACTIVATED',
                title='Membership Activated',
                body=f'Your {plan.name} membership has been activated by an administrator. Enjoy your premium benefits!',
                link_url='/settings/payments',
            )

            return ApiResponse(message=f"Plan '{plan.name}' activated directly for user.")
            
        elif action == 'extend':
            if not duration_days:
                return ApiResponse(success=False, message="duration_days is required for extension.", status=status.HTTP_400_BAD_REQUEST)
            
            membership = MemberMembership.objects.filter(member=target_user, is_active=True).first()
            if not membership:
                return ApiResponse(success=False, message="User has no active membership to extend.", status=status.HTTP_400_BAD_REQUEST)
            
            days = int(duration_days)
            if membership.end_date:
                membership.end_date += timedelta(days=days)
            else:
                membership.end_date = now + timedelta(days=days)
            
            membership.save(update_fields=['end_date'])
            
            audit(
                request, actor, action='MEMBERSHIP_DIRECT_EXTENDED', module='memberships',
                target_type='MEMBER', target_id=target_user.pk,
                new_data={'extended_days': days, 'new_end_date': str(membership.end_date)}
            )

            create_notification(
                recipient=target_user,
                type='MEMBERSHIP_EXTENDED',
                title='Membership Extended',
                body=f'Your membership has been extended by {days} days by an administrator. New expiry: {membership.end_date.strftime("%B %d, %Y") if membership.end_date else "lifetime"}.',
                link_url='/settings/payments',
            )

            return ApiResponse(message=f"Membership extended by {days} days.")
            
        elif action == 'cancel':
            membership = MemberMembership.objects.filter(member=target_user, is_active=True).first()
            if membership:
                membership.is_active = False
                membership.save(update_fields=['is_active'])
            
            target_user.is_premium = False
            target_user.save(update_fields=['is_premium', 'updated_at'])
            
            audit(
                request, actor, action='MEMBERSHIP_DIRECT_CANCELLED', module='memberships',
                target_type='MEMBER', target_id=target_user.pk,
                new_data={}
            )

            create_notification(
                recipient=target_user,
                type='MEMBERSHIP_CANCELLED',
                title='Membership Cancelled',
                body='Your membership has been cancelled by an administrator. Your premium benefits have ended.',
                link_url='/settings/payments',
            )

            return ApiResponse(message="Membership cancelled successfully.")
            
        else:
            return ApiResponse(success=False, message="Invalid action.", status=status.HTTP_400_BAD_REQUEST)

class _EnquiryWireSerializer(serializers.Serializer):
    def to_representation(self, enquiry):
        return {
            'id': str(enquiry.pk),
            'user': member_summary(enquiry.member, include_contact=True),
            'name': enquiry.name,
            'email': enquiry.email,
            'phone': enquiry.phone,
            'subject': enquiry.subject,
            'message': enquiry.message,
            'status': enquiry.status,
            'assigned_to': None,
            'internal_notes': enquiry.internal_notes,
            'created_at': enquiry.created_at,
            'updated_at': enquiry.updated_at,
        }

class AdminEnquiryListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'tickets.view_all'

    def get(self, request):
        queryset = ContactEnquiry.objects.select_related('member')
        requested_status = request.query_params.get('status')
        if requested_status:
            queryset = queryset.filter(status=requested_status)
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(email__icontains=search) | Q(subject__icontains=search)
            )
        return paginated_response(request, queryset, _EnquiryWireSerializer)

class AdminEnquiryDetailView(AdminEnquiryListView):
    @transaction.atomic
    def patch(self, request, enquiry_id):
        enquiry = get_object_or_404(ContactEnquiry.objects.select_for_update(), pk=enquiry_id)
        old = {'status': enquiry.status, 'internal_notes': enquiry.internal_notes}
        if 'status' in request.data:
            if request.data['status'] not in dict(ContactEnquiry.Status.choices):
                return bad_request('Invalid enquiry status.')
            enquiry.status = request.data['status']
            if enquiry.status == ContactEnquiry.Status.CONTACTED:
                enquiry.contacted_at = timezone.now()
            if enquiry.status == ContactEnquiry.Status.RESOLVED:
                enquiry.resolved_at = timezone.now()
        if 'internal_notes' in request.data:
            enquiry.internal_notes = str(request.data['internal_notes'])
        enquiry.save()
        audit(
            request, request.user, action='ENQUIRY_UPDATED', module='enquiries',
            target_type='CONTACT_ENQUIRY', target_id=enquiry.pk,
            old_data=old, new_data={'status': enquiry.status, 'internal_notes': enquiry.internal_notes},
        )
        return ApiResponse(data=_EnquiryWireSerializer(enquiry).data)

class _ComplaintWireSerializer(serializers.Serializer):
    def to_representation(self, complaint):
        return {
            'id': str(complaint.pk),
            'user': member_summary(complaint.member, include_contact=True),
            'subject': complaint.subject,
            'description': complaint.description,
            'status': complaint.status,
            'assigned_to': administrative_summary(complaint.assigned_to_staff),
            'escalated_by': administrative_summary(complaint.escalated_to_admin),
            'created_at': complaint.created_at,
            'updated_at': complaint.updated_at,
        }

class AdminComplaintListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'complaints.view_all'

    def get(self, request):
        queryset = Complaint.objects.select_related('member', 'assigned_to_staff', 'escalated_to_admin')
        requested_status = request.query_params.get('status')
        if requested_status:
            queryset = queryset.filter(status=requested_status)
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(Q(subject__icontains=search) | Q(description__icontains=search))
        return paginated_response(request, queryset, _ComplaintWireSerializer)

class AdminComplaintDetailView(AdminComplaintListView):
    required_permission = 'complaints.manage'

    @transaction.atomic
    def patch(self, request, complaint_id):
        complaint = get_object_or_404(Complaint.objects.select_for_update(), pk=complaint_id)
        new_status = request.data.get('status')
        if new_status not in dict(Complaint.Status.choices):
            return bad_request('Invalid complaint status.')
        old_status = complaint.status
        complaint.status = new_status
        if new_status == Complaint.Status.ESCALATED and str(request.user.account_type) == AccountType.ADMIN:
            complaint.escalated_to_admin = request.user
        complaint.save()
        audit(
            request, request.user, action=f'COMPLAINT_{new_status}', module='complaints',
            target_type='COMPLAINT', target_id=complaint.pk,
            old_data={'status': old_status}, new_data={'status': new_status},
        )
        return ApiResponse(data=_ComplaintWireSerializer(complaint).data)

class _ProfileReportWireSerializer(serializers.Serializer):
    def to_representation(self, report):
        return {
            'id': str(report.pk),
            'reported_user': member_summary(report.reported_member, include_contact=True),
            'reported_by': member_summary(report.reported_by_member, include_contact=True) if report.reported_by_member else None,
            'reason': report.reason,
            'details': report.details,
            'evidence_file': getattr(report, 'evidence_file', None),
            'status': report.status,
            'reviewed_by': administrative_summary(report.reviewed_by_admin or report.reviewed_by_staff),
            'created_at': report.created_at,
            'updated_at': report.updated_at,
        }

class AdminProfileReportListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'profile_reports.manage'

    def get(self, request):
        queryset = ProfileReport.objects.select_related(
            'reported_member', 'reported_by_member', 'reviewed_by_staff', 'reviewed_by_admin'
        )
        requested_status = request.query_params.get('status')
        if requested_status:
            queryset = queryset.filter(status=requested_status)
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(Q(reason__icontains=search) | Q(details__icontains=search))
        return paginated_response(request, queryset, _ProfileReportWireSerializer)

class AdminProfileReportDetailView(AdminProfileReportListView):
    @transaction.atomic
    def patch(self, request, report_id):
        report = get_object_or_404(ProfileReport.objects.select_for_update(), pk=report_id)
        new_status = request.data.get('status')
        if new_status not in dict(ProfileReport.Status.choices):
            return bad_request('Invalid report status.')
        old_status = report.status
        report.status = new_status
        if str(request.user.account_type) == AccountType.ADMIN:
            report.reviewed_by_admin = request.user
        report.save()
        audit(
            request, request.user, action=f'PROFILE_REPORT_{new_status}', module='safety',
            target_type='PROFILE_REPORT', target_id=report.pk,
            old_data={'status': old_status}, new_data={'status': new_status},
        )
        return ApiResponse(data=_ProfileReportWireSerializer(report).data)

class _SettingSerializer(serializers.ModelSerializer):
    class Meta:
        model = PlatformSetting
        fields = ('key', 'value', 'description', 'is_public', 'updated_at')
        read_only_fields = ('updated_at',)

class AdminSettingsView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)
    required_permission = 'settings.manage'

    def get(self, _request):
        return ApiResponse(data=_SettingSerializer(PlatformSetting.objects.all(), many=True).data)

    @transaction.atomic
    def patch(self, request):
        setting = get_object_or_404(PlatformSetting, key=request.data.get('key'))
        old = {'value': setting.value, 'description': setting.description, 'is_public': setting.is_public}
        serializer = _SettingSerializer(setting, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        setting = serializer.save(updated_by_super_admin=request.user)
        audit(
            request, request.user, action='SETTINGS_UPDATED', module='settings',
            target_type='PLATFORM_SETTING', target_id=setting.key,
            old_data=old, new_data={'value': setting.value, 'description': setting.description, 'is_public': setting.is_public},
        )
        return ApiResponse(data=_SettingSerializer(setting).data)

class _BackupSerializer(serializers.ModelSerializer):
    class Meta:
        model = BackupRecord
        fields = '__all__'

class AdminBackupListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)
    required_permission = 'backups.manage'

    def get(self, _request):
        return ApiResponse(data=_BackupSerializer(BackupRecord.objects.all(), many=True).data)

    def post(self, request):
        record = BackupRecord.objects.create(requested_by_super_admin=request.user)
        audit(
            request, request.user, action='BACKUP_REQUESTED', module='backups',
            target_type='BACKUP_RECORD', target_id=record.pk,
        )
        return ApiResponse(data=_BackupSerializer(record).data, status=status.HTTP_202_ACCEPTED)


class AdminSupportDashboardView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'tickets.view_all'

    def get(self, _request):
        today = timezone.localdate()
        return ApiResponse(data={
            'new_tickets': SupportTicket.objects.filter(status=SupportTicket.Status.OPEN).count(),
            'unassigned_tickets': SupportTicket.objects.filter(claimed_by_admin__isnull=True, claimed_by_super_admin__isnull=True).count(),
            'in_progress_tickets': SupportTicket.objects.filter(status=SupportTicket.Status.IN_PROGRESS).count(),
            'waiting_for_user': SupportTicket.objects.filter(status=SupportTicket.Status.WAITING_FOR_MEMBER).count(),
            'overdue_tickets': 0,
            'resolved_today': SupportTicket.objects.filter(resolved_at__date=today).count(),
            'closed_today': SupportTicket.objects.filter(closed_at__date=today).count(),
            'urgent_tickets': SupportTicket.objects.filter(priority=SupportTicket.Priority.URGENT).exclude(
                status__in=(SupportTicket.Status.RESOLVED, SupportTicket.Status.CLOSED)
            ).count(),
        })

class AdminSupportReportsView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'tickets.view_all'

    def get(self, _request):
        categories = dict(
            SupportTicket.objects.values_list('category__name').annotate(total=Count('id'))
        )
        return ApiResponse(data={
            'avg_first_response_minutes': 0,
            'avg_resolution_hours': 0,
            'ratings_distribution': {},
            'categories_breakdown': categories,
        })


class AdminFAQListCreateView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    def get(self, request):
        queryset = FAQ.objects.all()
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(question__icontains=search)
        from .api_utils import paginated_response
        return paginated_response(request, queryset.order_by('-created_at'), FAQSerializer)

    def post(self, request):
        serializer = FAQSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        item = serializer.save()
        return ApiResponse(data=FAQSerializer(item).data, status=status.HTTP_201_CREATED)


class AdminFAQDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    def get(self, _request, item_id):
        return ApiResponse(data=FAQSerializer(get_object_or_404(FAQ, pk=item_id)).data)

    def patch(self, request, item_id):
        item = get_object_or_404(FAQ, pk=item_id)
        serializer = FAQSerializer(item, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return ApiResponse(data=serializer.data)

    def delete(self, request, item_id):
        item = get_object_or_404(FAQ, pk=item_id)
        item.delete()
        return ApiResponse(message='FAQ deleted.')


class SuperAdminAccountListCreateView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)
    
    def get(self, request):
        from apps.accounts.models import Admin, Staff
        from django.db.models import Count, Q
        from django.utils import timezone
        search = request.query_params.get('search', '').strip()
        role_filter = request.query_params.get('role', '').upper()
        status_filter = request.query_params.get('status', '').lower()
        
        rows = []
        
        def filter_queryset(qs, is_support=False):
            if search:
                q_obj = Q(email__icontains=search) | Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(mobile_number__icontains=search)
                if is_support or hasattr(qs.model, 'employee_code'):
                    q_obj |= Q(employee_code__icontains=search)
                qs = qs.filter(q_obj)
            if status_filter == 'active':
                qs = qs.filter(is_active=True)
            elif status_filter == 'inactive':
                qs = qs.filter(is_active=False)
            return qs.filter(deleted_at__isnull=True)

        if not role_filter or role_filter == 'ADMIN':
            qs = Admin.objects.all()
            rows.extend(filter_queryset(qs))
        if not role_filter or role_filter == 'ADMIN':
            qs = Staff.objects.all().annotate(
                assigned_count=Count('work_assignments', filter=Q(work_assignments__status='ASSIGNED')),
                in_progress_count=Count('work_assignments', filter=Q(work_assignments__status='IN_PROGRESS')),
                completed_count=Count('work_assignments', filter=Q(work_assignments__status='COMPLETED')),
                overdue_count=Count('work_assignments', filter=Q(work_assignments__status='ASSIGNED', work_assignments__due_at__lt=timezone.now()))
            )
            rows.extend(filter_queryset(qs))

            
        rows.sort(key=lambda x: x.created_at, reverse=True)
        return paginated_response(request, rows, _AdministrativeAccountSerializer)

    @transaction.atomic
    def post(self, request):
        serializer = AdministrativeAccountCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        values = dict(serializer.validated_data)
        role = str(values.get('role', ''))
        account = _create_administrative_account(actor=request.user, values=values)
        
        audit(
            request, request.user,
            action=f'{role}_CREATED',
            module='accounts',
            target_type=role,
            target_id=account.pk,
            new_data={'email': account.email},
        )
        return ApiResponse(
            data=administrative_account_payload(account),
            message='Administrative account created.',
            status=status.HTTP_201_CREATED,
        )

class SuperAdminAccountDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)

    def _get_account(self, account_type, account_id):
        from apps.accounts.models import Admin, Staff, SuperAdmin
        model = {
            'admin': Admin,
            'staff': Staff,
            'super_admin': SuperAdmin,
            'super-admin': SuperAdmin,
            'ADMIN': Admin,
            'ADMIN': Staff,
            'SUPER_ADMIN': SuperAdmin,
        }.get(account_type.upper())
        if not model:
            raise PermissionDenied('Unsupported account type.')
        return get_object_or_404(model, pk=account_id)

    def get(self, request, account_type, account_id):
        account = self._get_account(account_type, account_id)
        return ApiResponse(data=administrative_account_payload(account))

    @transaction.atomic
    def patch(self, request, account_type, account_id):
        account = self._get_account(account_type, account_id)
        before = {'is_active': account.is_active, 'email': account.email}
        update_fields = []
        
        from apps.accounts.models import Admin, Staff, SuperAdmin, Member
        
        # 1. Uniqueness Validations
        email = request.data.get('email')
        if email:
            email = email.lower()
            if email != account.email:
                if (SuperAdmin.objects.filter(email__iexact=email).exists() or
                    Admin.objects.filter(email__iexact=email).exists() or
                    Staff.objects.filter(email__iexact=email).exists() or
                    Member.objects.filter(email__iexact=email).exists()):
                    return bad_request('This email is already registered in the system.', errors={'email': ['This email is already registered.']})
                account.email = email
                update_fields.append('email')

        mobile_number = request.data.get('mobile_number')
        if mobile_number:
            mobile_number = str(mobile_number).strip()
            if mobile_number != account.mobile_number:
                if (SuperAdmin.objects.filter(mobile_number=mobile_number).exists() or
                    Admin.objects.filter(mobile_number=mobile_number).exists() or
                    Staff.objects.filter(mobile_number=mobile_number).exists() or
                    Member.objects.filter(mobile_number=mobile_number).exists()):
                    return bad_request('This mobile number is already registered in the system.', errors={'mobile_number': ['This mobile number is already registered.']})
                account.mobile_number = mobile_number
                update_fields.append('mobile_number')

        employee_code = request.data.get('employee_code')
        if employee_code and hasattr(account, 'employee_code'):
            employee_code = str(employee_code).strip()
            if employee_code != account.employee_code:
                if (Admin.objects.filter(employee_code__iexact=employee_code).exists() or
                    Staff.objects.filter(employee_code__iexact=employee_code).exists()):
                    return bad_request('This employee code is already in use.', errors={'employee_code': ['This employee code is already in use.']})
                account.employee_code = employee_code
                update_fields.append('employee_code')

        # 2. Deactivation Safety checks
        if 'is_active' in request.data:
            is_active = bool(request.data['is_active'])
            if is_active != account.is_active:
                if not is_active:
                    from apps.core.models import WorkAssignment, SupportTicket
                    if str(account.account_type) == 'ADMIN':
                        active_tasks = WorkAssignment.objects.filter(
                            assigned_to_staff_id=account.pk, 
                            status__in=['ASSIGNED', 'IN_PROGRESS', 'WAITING', 'ESCALATED']
                        ).exists()
                        if active_tasks:
                            return bad_request('Cannot deactivate a Staff member with active assigned work.')

                account.is_active = is_active
                update_fields.append('is_active')

        # 3. Simple text/choice fields
        for field in ('first_name', 'last_name'):
            if field in request.data:
                setattr(account, field, request.data[field])
                update_fields.append(field)

        if 'support_level' in request.data and hasattr(account, 'support_level'):
            account.support_level = request.data['support_level']
            update_fields.append('support_level')

        if 'specialization' in request.data and hasattr(account, 'specialization'):
            account.specialization = request.data['specialization']
            update_fields.append('specialization')

        if 'manager_admin' in request.data and hasattr(account, 'manager_admin'):
            manager_val = request.data['manager_admin']
            if manager_val:
                account.manager_admin = Admin.objects.filter(pk=manager_val).first()
            else:
                account.manager_admin = None
            update_fields.append('manager_admin')

        if update_fields:
            account.save()

        audit(
            request, request.user,
            action='ACCOUNT_UPDATED',
            module='accounts',
            target_type=account_type.upper(),
            target_id=account.pk,
            old_data=before,
            new_data={'is_active': account.is_active, 'email': account.email},
        )
        return ApiResponse(data=administrative_account_payload(account), message='Account updated.')

    def delete(self, request, account_type, account_id):
        account = self._get_account(account_type, account_id)
        from apps.accounts.models import SuperAdmin, AccountType
        if str(account.account_type) == AccountType.SUPER_ADMIN:
            if SuperAdmin.objects.filter(deleted_at__isnull=True).count() <= 1:
                return bad_request('Cannot delete the final Super Admin.')
        
        if str(account.pk) == str(request.user.pk):
            return bad_request('Cannot delete your own currently logged-in account.')

        from apps.core.models import WorkAssignment
        if str(account.account_type) == 'ADMIN':
            active_tasks = WorkAssignment.objects.filter(
                assigned_to_staff_id=account.pk, 
                status__in=['ASSIGNED', 'IN_PROGRESS', 'WAITING', 'ESCALATED']
            ).exists()
            if active_tasks:
                return bad_request('Cannot delete a Staff member with active assigned work.')

        account.is_active = False
        account.deleted_at = timezone.now()
        account.token_version += 1
        account.save(update_fields=('is_active', 'deleted_at', 'token_version', 'updated_at'))
        audit(
            request, request.user,
            action='ACCOUNT_DELETED',
            module='accounts',
            target_type=account_type.upper(),
            target_id=account.pk,
        )
        return ApiResponse(message='Account safely archived and deleted.')

class SuperAdminAccountActionView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)

    def _get_account(self, account_type, account_id):
        from apps.accounts.models import Admin, Staff, SuperAdmin as SuperAdminModel
        model = {
            'admin': Admin,
            'staff': Staff,
            'ADMIN': Admin,
            'ADMIN': Staff,
            'super_admin': SuperAdminModel,
            'SUPER_ADMIN': SuperAdminModel,
        }.get(account_type.upper(), None) or {
            'admin': Admin,
            'staff': Staff,
            'super_admin': SuperAdminModel,
        }.get(account_type.lower())
        if model is None:
            from django.http import Http404
            raise Http404('Unknown account type.')
        return get_object_or_404(model, pk=account_id)

    def post(self, request, account_type, account_id, action):
        from apps.accounts.models import SuperAdmin as SuperAdminModel
        account = self._get_account(account_type, account_id)
        if action == 'activate':
            account.is_active = True
            account.save(update_fields=('is_active', 'updated_at'))
            return ApiResponse(message='Account activated.')
        elif action == 'deactivate':
            # Protect the last Super Admin
            if isinstance(account, SuperAdminModel):
                active_count = SuperAdminModel.objects.filter(
                    is_active=True, deleted_at__isnull=True
                ).count()
                if active_count <= 1:
                    return bad_request('Cannot deactivate the last active Super Admin.')
            # Deactivation safety for Staff and CSA
            if str(account.account_type) == AccountType.ADMIN:
                from apps.core.models import WorkAssignment
                if WorkAssignment.objects.filter(
                    assigned_to_staff_id=account.pk,
                    status__in=['ASSIGNED', 'IN_PROGRESS']
                ).exists():
                    return bad_request('Cannot deactivate a Staff member with active work assignments.')

            account.is_active = False
            account.token_version += 1
            account.save(update_fields=('is_active', 'token_version', 'updated_at'))
            return ApiResponse(message='Account deactivated.')
        elif action == 'reset-password':
            new_password = request.data.get('new_password')
            if not new_password:
                return bad_request('new_password is required.')
            try:
                validate_password(new_password, user=account)
            except DjangoValidationError as exc:
                return bad_request('Choose a stronger password.', errors={'new_password': exc.messages})
            account.set_password(new_password)
            account.password_changed_at = timezone.now()
            account.token_version += 1
            account.save(update_fields=('password', 'password_changed_at', 'token_version', 'updated_at'))
            return ApiResponse(message='Password reset successfully.')
        return bad_request('Unsupported action.')

class SuperAdminAccountActivityView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)
    
    def get(self, request, account_type, account_id):
        from apps.accounts.models import SuperAdminActivityLog, AdminActivityLog, StaffActivityLog
        logs = []
        for model in (SuperAdminActivityLog, AdminActivityLog, StaffActivityLog):
            logs.extend(model.objects.filter(actor_id=account_id))
        logs.sort(key=lambda x: x.created_at, reverse=True)
        data = [{
            'id': str(log.id),
            'actor_name': log.actor_name,
            'actor_role': log.actor_role,
            'action': log.action,
            'module': log.module,
            'description': log.description,
            'created_at': log.created_at,
            'ip_address': log.ip_address,
        } for log in logs[:100]]
        return ApiResponse(data=data)

class AdminStaffListCreateView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    def get(self, request):
        if not request.user.has_admin_permission('staff.view') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
            raise PermissionDenied('You do not have staff.view permission.')
        from apps.accounts.models import Staff
        from django.db.models import Count, Q
        from django.utils import timezone
        
        search = request.query_params.get('search', '').strip()
        status_filter = request.query_params.get('status', '').lower()
        
        queryset = Staff.objects.filter(deleted_at__isnull=True).annotate(
            assigned_count=Count('work_assignments', filter=Q(work_assignments__status='ASSIGNED')),
            in_progress_count=Count('work_assignments', filter=Q(work_assignments__status='IN_PROGRESS')),
            completed_count=Count('work_assignments', filter=Q(work_assignments__status='COMPLETED')),
            overdue_count=Count('work_assignments', filter=Q(work_assignments__status='ASSIGNED', work_assignments__due_at__lt=timezone.now()))
        )
        if search:
            queryset = queryset.filter(
                Q(email__icontains=search) | Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(mobile_number__icontains=search) | Q(employee_code__icontains=search)
            )
        if status_filter == 'active':
            queryset = queryset.filter(is_active=True)
        elif status_filter == 'inactive':
            queryset = queryset.filter(is_active=False)
            
        return paginated_response(request, queryset.order_by('-created_at'), _AdministrativeAccountSerializer)

    @transaction.atomic
    def post(self, request):
        if not request.user.has_admin_permission('staff.create') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
            raise PermissionDenied('You do not have staff.create permission.')
        serializer = AdministrativeAccountCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        values = dict(serializer.validated_data)
        values['role'] = RoleCode.ADMIN
        account = _create_administrative_account(actor=request.user, values=values)
        
        audit(
            request, request.user,
            action='STAFF_CREATED',
            module='accounts',
            target_type='ADMIN',
            target_id=account.pk,
            new_data={'email': account.email},
        )
        return ApiResponse(
            data=administrative_account_payload(account),
            message='Staff account created.',
            status=status.HTTP_201_CREATED,
        )

class AdminStaffDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    def get(self, request, pk):
        if not request.user.has_admin_permission('staff.view') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
            raise PermissionDenied('You do not have staff.view permission.')
        from apps.accounts.models import Staff
        obj = get_object_or_404(Staff, pk=pk, deleted_at__isnull=True)
        return ApiResponse(data=administrative_account_payload(obj))
        
    @transaction.atomic
    def patch(self, request, pk):
        if not request.user.has_admin_permission('staff.manage') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
            raise PermissionDenied('You do not have staff.manage permission.')
        from apps.accounts.models import Staff, SuperAdmin, Admin, Member
        obj = get_object_or_404(Staff, pk=pk, deleted_at__isnull=True)
        update_fields = []
        
        # 1. Uniqueness Validations
        email = request.data.get('email')
        if email:
            email = email.lower()
            if email != obj.email:
                if (SuperAdmin.objects.filter(email__iexact=email).exists() or
                    Admin.objects.filter(email__iexact=email).exists() or
                    Staff.objects.filter(email__iexact=email).exists() or
                    Member.objects.filter(email__iexact=email).exists()):
                    return bad_request('This email is already registered in the system.', errors={'email': ['This email is already registered.']})
                obj.email = email
                update_fields.append('email')

        mobile_number = request.data.get('mobile_number')
        if mobile_number:
            mobile_number = str(mobile_number).strip()
            if mobile_number != obj.mobile_number:
                if (SuperAdmin.objects.filter(mobile_number=mobile_number).exists() or
                    Admin.objects.filter(mobile_number=mobile_number).exists() or
                    Staff.objects.filter(mobile_number=mobile_number).exists() or
                    Member.objects.filter(mobile_number=mobile_number).exists()):
                    return bad_request('This mobile number is already registered in the system.', errors={'mobile_number': ['This mobile number is already registered.']})
                obj.mobile_number = mobile_number
                update_fields.append('mobile_number')

        employee_code = request.data.get('employee_code')
        if employee_code:
            employee_code = str(employee_code).strip()
            if employee_code != obj.employee_code:
                if (Admin.objects.filter(employee_code__iexact=employee_code).exists() or
                    Staff.objects.filter(employee_code__iexact=employee_code).exists()):
                    return bad_request('This employee code is already in use.', errors={'employee_code': ['This employee code is already in use.']})
                obj.employee_code = employee_code
                update_fields.append('employee_code')

        # 2. Deactivation Safety checks
        if 'is_active' in request.data:
            is_active = bool(request.data['is_active'])
            if is_active != obj.is_active:
                if not is_active:
                    from apps.core.models import WorkAssignment
                    active_tasks = WorkAssignment.objects.filter(
                        assigned_to_staff_id=obj.pk, 
                        status__in=['ASSIGNED', 'IN_PROGRESS', 'WAITING', 'ESCALATED']
                    ).exists()
                    if active_tasks:
                        return bad_request('Cannot deactivate a Staff member with active assigned work.')
                obj.is_active = is_active
                update_fields.append('is_active')

        # 3. Simple text fields
        for field in ('first_name', 'last_name'):
            if field in request.data:
                setattr(obj, field, request.data[field])
                update_fields.append(field)

        if update_fields:
            obj.save()
            
        audit(
            request, request.user,
            action='STAFF_UPDATED',
            module='accounts',
            target_type='ADMIN',
            target_id=obj.pk,
        )
        return ApiResponse(data=administrative_account_payload(obj), message='Staff details updated.')

class AdminStaffActionView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    def post(self, request, pk, action):
        from apps.accounts.models import Staff
        obj = get_object_or_404(Staff, pk=pk, deleted_at__isnull=True)
        if action == 'activate':
            if not request.user.has_admin_permission('staff.manage') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
                raise PermissionDenied('You do not have staff.manage permission.')
            obj.is_active = True
            obj.save(update_fields=('is_active', 'updated_at'))
            return ApiResponse(message='Staff activated.')
        elif action == 'deactivate':
            if not request.user.has_admin_permission('staff.manage') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
                raise PermissionDenied('You do not have staff.manage permission.')
            from apps.core.models import WorkAssignment
            active_tasks = WorkAssignment.objects.filter(
                assigned_to_staff_id=obj.pk, 
                status__in=['ASSIGNED', 'IN_PROGRESS', 'WAITING', 'ESCALATED']
            ).exists()
            if active_tasks:
                return bad_request('Cannot deactivate a Staff member with active assigned work.')
            obj.is_active = False
            obj.token_version += 1
            obj.save(update_fields=('is_active', 'token_version', 'updated_at'))
            return ApiResponse(message='Staff deactivated.')
        elif action == 'reset-password':
            if not request.user.has_admin_permission('staff.manage') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
                raise PermissionDenied('You do not have staff.manage permission.')
            new_password = request.data.get('new_password')
            if not new_password:
                return bad_request('new_password is required.')
            try:
                validate_password(new_password, user=obj)
            except DjangoValidationError as exc:
                return bad_request('Choose a stronger password.', errors={'new_password': exc.messages})
            obj.set_password(new_password)
            obj.password_changed_at = timezone.now()
            obj.token_version += 1
            obj.save(update_fields=('password', 'password_changed_at', 'token_version', 'updated_at'))
            return ApiResponse(message='Password reset successfully.')
        return bad_request('Unsupported action.')

class AdminStaffActivityView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    def get(self, request, pk):
        if not request.user.has_admin_permission('staff.activity') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
            raise PermissionDenied('You do not have staff.activity permission.')
        from apps.accounts.models import StaffActivityLog
        logs = StaffActivityLog.objects.filter(actor_id=pk).order_by('-created_at')[:100]
        data = [{
            'id': str(log.id),
            'actor_name': log.actor_name,
            'actor_role': log.actor_role,
            'action': log.action,
            'module': log.module,
            'description': log.description,
            'created_at': log.created_at,
            'ip_address': log.ip_address,
        } for log in logs]
        return ApiResponse(data=data)



class WorkAssignmentSerializer(serializers.ModelSerializer):
    assigned_to_staff_details = _AdministrativeAccountSerializer(source='assigned_to_staff', read_only=True)
    assignment_type_display = serializers.CharField(source='get_assignment_type_display', read_only=True)
    assigned_by_name = serializers.SerializerMethodField()
    member_name = serializers.SerializerMethodField()
    member_email = serializers.SerializerMethodField()
    profile_photos = serializers.SerializerMethodField()

    class Meta:
        from apps.core.models import WorkAssignment
        model = WorkAssignment
        fields = '__all__'

    @staticmethod
    def _member(obj):
        if obj.related_profile_verification_id:
            return obj.related_profile_verification.member
        if obj.related_document_verification_id:
            return obj.related_document_verification.member
        if obj.related_complaint_id:
            return obj.related_complaint.member
        if obj.related_profile_report_id:
            return obj.related_profile_report.reported_member
        return None

    def get_assigned_by_name(self, obj):
        actor = obj.assigned_by_admin or obj.assigned_by_super_admin
        return actor.get_full_name() if actor else ''

    def get_member_name(self, obj):
        member = self._member(obj)
        return member.get_full_name() if member else ''

    def get_member_email(self, obj):
        member = self._member(obj)
        return member.email if member else ''

    def get_profile_photos(self, obj):
        request_record = obj.related_profile_verification
        if (
            obj.assignment_type != 'PHOTO_VERIFICATION'
            or request_record is None
            or request_record.verification_type != ProfileVerificationRequest.VerificationType.PROFILE_PHOTO
        ):
            return []
        from apps.profiles.models import ProfilePhoto
        from apps.profiles.serializers import ProfilePhotoSerializer

        photos = (
            ProfilePhoto.objects.without_binary()
            .filter(user_id=request_record.member_id, status=ProfilePhoto.Status.PENDING)
            .order_by('display_order', 'created_at')
        )
        return ProfilePhotoSerializer(photos, many=True, context=self.context).data

class AdminAssignWorkView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    @transaction.atomic
    def post(self, request):
        if not request.user.has_admin_permission('verification.assign') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
            raise PermissionDenied('You do not have verification.assign permission.')
        from apps.accounts.models import Staff
        from apps.core.models import WorkAssignment, ProfileVerificationRequest, Complaint, ProfileReport, ProfileVerificationAssignment, Notification
        from apps.accounts.models import MemberDocument
        
        staff_id = request.data.get('assigned_to_staff')
        assignment_type = request.data.get('assignment_type')
        priority = request.data.get('priority', 'NORMAL')
        due_at = request.data.get('due_at')
        notes = request.data.get('notes', '')
        related_id = request.data.get('related_id')

        if not staff_id or not assignment_type:
            return bad_request('assigned_to_staff and assignment_type are required.')

        import uuid as _uuid
        try:
            _uuid.UUID(str(staff_id))
        except (ValueError, AttributeError):
            return bad_request(f'assigned_to_staff "{staff_id}" is not a valid identifier.')
        if related_id:
            try:
                _uuid.UUID(str(related_id))
            except (ValueError, AttributeError):
                return bad_request(f'related_id "{related_id}" is not a valid identifier.')

        staff = get_object_or_404(Staff, pk=staff_id, deleted_at__isnull=True)
        if not staff.is_active:
            return bad_request('Cannot assign work to an inactive staff member.')
            
        admin = request.user if str(request.user.account_type) == AccountType.ADMIN else None
        super_admin = request.user if str(request.user.account_type) == AccountType.SUPER_ADMIN else None
        
        assignment = WorkAssignment(
            assignment_type=assignment_type,
            assigned_to_staff=staff,
            assigned_by_admin=admin,
            assigned_by_super_admin=super_admin,
            priority=priority,
            status='ASSIGNED',
            due_at=due_at,
            notes=notes,
        )
        
        # Decrement previous assignee workload if applicable
        if assignment_type in ('PROFILE_VERIFICATION', 'PHOTO_VERIFICATION') and related_id:
            previous_assignment = ProfileVerificationAssignment.objects.filter(verification_request_id=related_id, is_current=True).first()
            if previous_assignment and previous_assignment.assigned_to_staff:
                prev_workload = getattr(previous_assignment.assigned_to_staff, 'workload', None)
                if prev_workload:
                    prev_workload.open_verifications_count = max(0, prev_workload.open_verifications_count - 1)
                    prev_workload.current_workload_score = prev_workload.open_verifications_count + prev_workload.open_tickets_count
                    prev_workload.save()

        if assignment_type in ('PROFILE_VERIFICATION', 'PHOTO_VERIFICATION'):
            req = get_object_or_404(ProfileVerificationRequest, pk=related_id)
            req.status = 'ASSIGNED'
            req.save()
            
            ProfileVerificationAssignment.objects.filter(verification_request=req, is_current=True).update(is_current=False)
            ProfileVerificationAssignment.objects.create(
                verification_request=req,
                assigned_to_staff=staff,
                assigned_by_admin=admin,
                assigned_by_super_admin=super_admin,
                is_current=True
            )
            assignment.related_profile_verification = req
            
        elif assignment_type == 'DOCUMENT_VERIFICATION':
            doc = get_object_or_404(MemberDocument, pk=related_id)
            assignment.related_document_verification = doc
            
        elif assignment_type == 'COMPLAINT_REVIEW':
            comp = get_object_or_404(Complaint, pk=related_id)
            comp.status = 'UNDER_REVIEW'
            comp.assigned_to_staff = staff
            comp.save()
            assignment.related_complaint = comp
            
        elif assignment_type == 'PROFILE_REPORT_REVIEW':
            rep = get_object_or_404(ProfileReport, pk=related_id)
            rep.status = 'UNDER_REVIEW'
            rep.reviewed_by_staff = staff
            rep.save()
            assignment.related_profile_report = rep
            
        assignment.save()

        # Increment new assignee workload
        from apps.core.models import Workload
        new_workload, _ = Workload.objects.get_or_create(staff_member=staff)
        new_workload.open_verifications_count += 1
        new_workload.current_workload_score = new_workload.open_verifications_count + new_workload.open_tickets_count
        new_workload.save()
        
        Notification.objects.create(
            staff_recipient=staff,
            notification_type='WORK_ASSIGNMENT',
            title='New task assigned',
            message=f'You have been assigned a new task: {assignment.get_assignment_type_display()}. Due by: {due_at or "N/A"}. Notes: {notes}',
            related_object_type='WorkAssignment',
            related_object_id=str(assignment.id),
        )
        
        return ApiResponse(data=WorkAssignmentSerializer(assignment).data, message='Work assigned successfully.')

class AdminAssignTicketView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    @transaction.atomic
    def post(self, request):
        return bad_request('Direct assignment endpoint retired. Use claim_ticket or AdminTicketActionView instead.')
            


class StaffWorkListView(ScopedAPIView):
    allowed_account_types = (AccountType.ADMIN,)
    
    def get(self, request):
        from apps.core.models import WorkAssignment
        from django.db.models import Case, When, Value, IntegerField
        
        queryset = WorkAssignment.objects.filter(assigned_to_staff=request.user).select_related(
            'assigned_by_admin',
            'assigned_by_super_admin',
            'related_profile_verification__member',
            'related_document_verification__member',
            'related_complaint__member',
            'related_profile_report__reported_member',
        )
        
        search = request.query_params.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(related_profile_verification__member__first_name__icontains=search) |
                Q(related_profile_verification__member__last_name__icontains=search) |
                Q(related_profile_verification__member__email__icontains=search) |
                Q(related_document_verification__member__first_name__icontains=search) |
                Q(related_document_verification__member__last_name__icontains=search) |
                Q(related_document_verification__member__email__icontains=search) |
                Q(related_complaint__member__first_name__icontains=search) |
                Q(related_complaint__member__last_name__icontains=search) |
                Q(related_complaint__member__email__icontains=search) |
                Q(related_profile_report__reported_member__first_name__icontains=search) |
                Q(related_profile_report__reported_member__last_name__icontains=search) |
                Q(related_profile_report__reported_member__email__icontains=search)
            )

        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter.upper())
            
        type_filter = request.query_params.get('assignment_type')
        if type_filter:
            queryset = queryset.filter(assignment_type=type_filter.upper())
            
        priority_filter = request.query_params.get('priority')
        if priority_filter:
            queryset = queryset.filter(priority=priority_filter.upper())

        due_filter = request.query_params.get('due_date')
        if due_filter == 'today':
            queryset = queryset.filter(due_at__date=timezone.now().date())
        elif due_filter == 'overdue':
            queryset = queryset.filter(due_at__lt=timezone.now()).exclude(status='COMPLETED')

        ordering = request.query_params.get('ordering', '-created_at')
        if ordering == 'priority':
            priority_order = Case(
                When(priority='URGENT', then=Value(3)),
                When(priority='HIGH', then=Value(2)),
                When(priority='NORMAL', then=Value(1)),
                When(priority='LOW', then=Value(0)),
                default=Value(0),
                output_field=IntegerField(),
            )
            queryset = queryset.annotate(p_val=priority_order).order_by('-p_val')
        elif ordering == 'due_date':
            queryset = queryset.order_by('due_at')
        else:
            queryset = queryset.order_by('-created_at')
            
        return paginated_response(request, queryset, WorkAssignmentSerializer)

class StaffWorkActionView(ScopedAPIView):
    allowed_account_types = (AccountType.ADMIN,)
    
    @transaction.atomic
    def post(self, request):
        from apps.core.models import WorkAssignment, ProfileVerificationHistory
        assignment_id = request.data.get('assignment_id')
        action = request.data.get('action')
        notes = request.data.get('notes', '')
        outcome = request.data.get('outcome', '').upper()
        
        assignment = get_object_or_404(WorkAssignment, pk=assignment_id, assigned_to_staff=request.user)
        
        if action == 'start':
            assignment.status = 'IN_PROGRESS'
            assignment.started_at = timezone.now()
            assignment.save()
            
            if assignment.related_profile_verification:
                req = assignment.related_profile_verification
                req.status = 'IN_REVIEW'
                req.save()
            elif assignment.related_complaint:
                comp = assignment.related_complaint
                comp.status = 'UNDER_REVIEW'
                comp.save()
                
            return ApiResponse(message='Task started.')
            
        elif action == 'complete':
            request_record = assignment.related_profile_verification
            if (
                request_record
                and request_record.verification_type
                == ProfileVerificationRequest.VerificationType.PROFILE_PHOTO
            ):
                from apps.profiles.models import ProfilePhoto

                if ProfilePhoto.objects.filter(
                    user_id=request_record.member_id,
                    status=ProfilePhoto.Status.PENDING,
                ).exists():
                    return bad_request(
                        'Review each pending photo with its approve or reject action before completing this task.'
                    )
                assignment.status = 'COMPLETED'
                assignment.completed_at = timezone.now()
                assignment.save(update_fields=('status', 'completed_at', 'updated_at'))
                return ApiResponse(message='Photo verification task completed.')

            assignment.status = 'COMPLETED'
            assignment.completed_at = timezone.now()
            assignment.save()
            
            if assignment.related_profile_verification:
                req = assignment.related_profile_verification
                old_status = req.status
                new_status = 'APPROVED' if outcome != 'REJECT' else 'REJECTED'
                req.status = new_status
                req.reviewed_at = timezone.now()
                if new_status == 'APPROVED':
                    req.approved_at = timezone.now()
                    member = req.member
                    if req.verification_type == 'FULL_PROFILE':
                        member.profile_status = Member.ProfileStatus.APPROVED
                    elif req.verification_type == 'IDENTITY_DOCUMENT':
                        member.document_status = Member.DocumentStatus.APPROVED
                        member.documents.filter(status='PENDING').update(
                            status='APPROVED',
                            reviewed_at=timezone.now(),
                            reviewed_by_id=request.user.pk
                        )
                    member.save()
                else:
                    req.rejected_at = timezone.now()
                    req.rejection_reason = notes
                    member = req.member
                    if req.verification_type == 'FULL_PROFILE':
                        member.profile_status = Member.ProfileStatus.REJECTED
                    elif req.verification_type == 'IDENTITY_DOCUMENT':
                        member.document_status = Member.DocumentStatus.REJECTED
                        member.documents.filter(status='PENDING').update(
                            status='REJECTED',
                            rejection_reason=notes,
                            reviewed_at=timezone.now(),
                            reviewed_by_id=request.user.pk
                        )
                    member.save()
                req.save()
                
                ProfileVerificationHistory.objects.create(
                    verification_request=req,
                    old_status=old_status,
                    new_status=new_status,
                    changed_by_staff=request.user,
                    reason=notes
                )
                
            elif assignment.related_complaint:
                comp = assignment.related_complaint
                comp.status = 'RESOLVED'
                comp.save()
                
            elif assignment.related_profile_report:
                rep = assignment.related_profile_report
                rep.status = 'ACTIONED' if outcome == 'APPROVE' else 'DISMISSED'
                rep.save()
                
            return ApiResponse(message='Task completed.')
            
        elif action == 'escalate':
            assignment.status = 'ESCALATED'
            assignment.save()
            
            if assignment.related_profile_verification:
                req = assignment.related_profile_verification
                old_status = req.status
                req.status = 'ESCALATED'
                req.escalation_reason = notes
                req.save()
                
                ProfileVerificationHistory.objects.create(
                    verification_request=req,
                    old_status=old_status,
                    new_status='ESCALATED',
                    changed_by_staff=request.user,
                    reason=notes
                )
                
            elif assignment.related_complaint:
                comp = assignment.related_complaint
                comp.status = 'ESCALATED'
                if request.user.manager_admin:
                    comp.escalated_to_admin = request.user.manager_admin
                comp.save()
                
            return ApiResponse(message='Task escalated to admin.')
            
        return bad_request('Invalid action.')

class AdminDuplicateFlagListView(ScopedAPIView):
    """
    GET  /api/admin/duplicate-flags/         — list all duplicate flags
    PATCH /api/admin/duplicate-flags/<pk>/   — update review_status + note

    Accessible by: Super Admin, Admin, Staff
    """
    allowed_account_types = (
        AccountType.SUPER_ADMIN,
        AccountType.ADMIN,
        AccountType.ADMIN,
    )

    def get(self, request):
        from apps.accounts.models import DuplicateAccountFlag

        queryset = DuplicateAccountFlag.objects.select_related(
            'primary_member',
            'duplicate_member',
            'reviewed_by_admin',
            'reviewed_by_super_admin',
        ).order_by('-created_at')

        # Filter by review status
        review_status = request.query_params.get('status', '').strip().upper()
        if review_status:
            queryset = queryset.filter(review_status=review_status)

        # Filter by flag type
        flag_type = request.query_params.get('flag_type', '').strip().upper()
        if flag_type:
            queryset = queryset.filter(flag_type=flag_type)

        # Filter by member (search email)
        member_email = request.query_params.get('email', '').strip()
        if member_email:
            queryset = queryset.filter(
                Q(primary_member__email__icontains=member_email)
                | Q(duplicate_member__email__icontains=member_email)
            )

        def _flag_payload(flag):
            def _member_summary(m):
                return {
                    'id': str(m.pk),
                    'email': m.email,
                    'full_name': m.get_full_name(),
                    'profile_status': m.profile_status,
                    'is_active': m.is_active,
                }
            return {
                'id': str(flag.pk),
                'flag_type': flag.flag_type,
                'review_status': flag.review_status,
                'similarity_score': flag.similarity_score,
                'match_detail': flag.match_detail,
                'primary_member': _member_summary(flag.primary_member),
                'duplicate_member': _member_summary(flag.duplicate_member),
                'review_note': flag.review_note,
                'reviewed_at': flag.reviewed_at,
                'auto_detected': flag.auto_detected,
                'created_at': flag.created_at,
            }

        # Manual pagination
        import math
        from django.core.paginator import Paginator, EmptyPage
        try:
            page_size = max(1, min(int(request.query_params.get('page_size', 20)), 100))
        except (TypeError, ValueError):
            page_size = 20
        try:
            page_num = max(1, int(request.query_params.get('page', 1)))
        except (TypeError, ValueError):
            page_num = 1

        paginator = Paginator(queryset, page_size)
        try:
            page = paginator.page(page_num)
        except EmptyPage:
            page = paginator.page(paginator.num_pages or 1)

        results = [_flag_payload(f) for f in page.object_list]
        return ApiResponse(data={
            'count': paginator.count,
            'page': page.number,
            'page_size': page_size,
            'num_pages': max(1, math.ceil(paginator.count / page_size)),
            'next': page.next_page_number() if page.has_next() else None,
            'previous': page.previous_page_number() if page.has_previous() else None,
            'results': results,
        })

    @transaction.atomic
    def patch(self, request, pk):
        from apps.accounts.models import DuplicateAccountFlag

        flag = get_object_or_404(DuplicateAccountFlag, pk=pk)
        new_status = request.data.get('review_status', '').strip().upper()
        valid_statuses = [s[0] for s in DuplicateAccountFlag.ReviewStatus.choices]
        if new_status not in valid_statuses:
            return bad_request(f'review_status must be one of: {", ".join(valid_statuses)}.')

        flag.review_status = new_status
        flag.review_note = str(request.data.get('review_note', flag.review_note)).strip()
        flag.reviewed_at = timezone.now()

        account_type = getattr(request.user, 'account_type', None)
        if account_type == AccountType.SUPER_ADMIN:
            flag.reviewed_by_super_admin = request.user
        elif account_type in (AccountType.ADMIN, AccountType.ADMIN):
            flag.reviewed_by_admin = getattr(request.user, 'manager_admin', None) or (
                request.user if account_type == AccountType.ADMIN else None
            )

        flag.save()
        return ApiResponse(message=f'Duplicate flag updated to {new_status}.')

class AdminEligibleStaffListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    def get(self, request):
        from apps.accounts.models import Staff
        
        queryset = Staff.objects.filter(is_active=True, deleted_at__isnull=True).order_by('first_name', 'last_name')
        
        data = []
        for s in queryset:
            is_online = s.availability.is_online if hasattr(s, 'availability') else True
            availability_status = s.availability.availability_status if hasattr(s, 'availability') else 'AVAILABLE'
            is_suspended = s.availability.is_suspended if hasattr(s, 'availability') else False
            last_active = s.availability.last_active_at if hasattr(s, 'availability') else None
            
            assigned = s.workload.open_verifications_count if hasattr(s, 'workload') else 0
            capacity = s.workload.capacity if hasattr(s, 'workload') else 10
            current_score = s.workload.current_workload_score if hasattr(s, 'workload') else 0
            
            data.append({
                'id': str(s.id),
                'full_name': s.get_full_name(),
                'email': s.email,
                'employee_code': s.employee_code,
                'is_online': is_online,
                'availability_status': availability_status,
                'is_suspended': is_suspended,
                'last_active_at': last_active,
                'workload': {
                    'assigned': assigned,
                    'capacity': capacity,
                    'current_score': current_score,
                }
            })
        return ApiResponse(data=data)

class AdminEligibleAgentsListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    def get(self, request):
        return ApiResponse(data=[])

from rest_framework import viewsets, permissions
from apps.core.models import Specialization, Queue, AssignmentStrategy, EmployeeAvailability, Workload, AssignmentRule, AssignmentAudit

class SpecializationSerializer(serializers.ModelSerializer):
    class Meta:
        model = Specialization
        fields = '__all__'

class QueueSerializer(serializers.ModelSerializer):
    class Meta:
        model = Queue
        fields = '__all__'

class AssignmentStrategySerializer(serializers.ModelSerializer):
    class Meta:
        model = AssignmentStrategy
        fields = '__all__'

class EmployeeAvailabilitySerializer(serializers.ModelSerializer):
    class Meta:
        model = EmployeeAvailability
        fields = '__all__'

class WorkloadSerializer(serializers.ModelSerializer):
    class Meta:
        model = Workload
        fields = '__all__'

class AssignmentRuleSerializer(serializers.ModelSerializer):
    class Meta:
        model = AssignmentRule
        fields = '__all__'

class AssignmentAuditSerializer(serializers.ModelSerializer):
    class Meta:
        model = AssignmentAudit
        fields = '__all__'

class AdminBulkReassignTicketView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    @transaction.atomic
    def post(self, request):
        return bad_request('Customer support agents no longer supported.')

class AdminBulkReassignWorkView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    @transaction.atomic
    def post(self, request):
        if not request.user.has_admin_permission('verification.assign') and str(request.user.account_type) != AccountType.SUPER_ADMIN:
            raise PermissionDenied('You do not have verification.assign permission.')
        from apps.accounts.models import Staff
        from apps.core.models import WorkAssignment, ProfileVerificationRequest, ProfileVerificationAssignment, Workload, Notification
        
        verification_ids = request.data.get('verification_ids', [])
        staff_id = request.data.get('assigned_to_staff')
        notes = request.data.get('notes', 'Bulk reassigned')
        
        if not verification_ids or not staff_id:
            return bad_request('Both verification_ids and assigned_to_staff are required.')
            
        staff = get_object_or_404(Staff, pk=staff_id, deleted_at__isnull=True)
        if not staff.is_active:
            return bad_request('Cannot assign tasks to an inactive staff member.')
            
        admin = request.user if str(request.user.account_type) == AccountType.ADMIN else None
        super_admin = request.user if str(request.user.account_type) == AccountType.SUPER_ADMIN else None
        
        assigned_count = 0
        for vid in verification_ids:
            try:
                req = ProfileVerificationRequest.objects.get(pk=vid)
            except ProfileVerificationRequest.DoesNotExist:
                continue
                
            previous_assignment = ProfileVerificationAssignment.objects.filter(verification_request=req, is_current=True).first()
            if previous_assignment and previous_assignment.assigned_to_staff:
                prev_workload = getattr(previous_assignment.assigned_to_staff, 'workload', None)
                if prev_workload:
                    prev_workload.open_verifications_count = max(0, prev_workload.open_verifications_count - 1)
                    prev_workload.current_workload_score = prev_workload.open_verifications_count + prev_workload.open_tickets_count
                    prev_workload.save()
            
            req.status = 'ASSIGNED'
            req.save()
            
            ProfileVerificationAssignment.objects.filter(verification_request=req, is_current=True).update(is_current=False)
            ProfileVerificationAssignment.objects.create(
                verification_request=req,
                assigned_to_staff=staff,
                assigned_by_admin=admin,
                assigned_by_super_admin=super_admin,
                is_current=True
            )
            
            WorkAssignment.objects.filter(related_profile_verification=req).delete()
            
            assignment_type = 'PROFILE_VERIFICATION'
            if req.verification_type == 'PROFILE_PHOTO':
                assignment_type = 'PHOTO_VERIFICATION'
            elif req.verification_type == 'IDENTITY_DOCUMENT':
                assignment_type = 'DOCUMENT_VERIFICATION'
                
            WorkAssignment.objects.create(
                assignment_type=assignment_type,
                assigned_to_staff=staff,
                assigned_by_admin=admin,
                assigned_by_super_admin=super_admin,
                related_profile_verification=req,
                priority=req.priority,
                status='ASSIGNED',
                due_at=timezone.now() + timezone.timedelta(days=1),
                notes=notes
            )
            
            assigned_count += 1
            
        if assigned_count > 0:
            new_workload, _ = Workload.objects.get_or_create(staff_member=staff)
            new_workload.open_verifications_count = ProfileVerificationAssignment.objects.filter(assigned_to_staff=staff, is_current=True, verification_request__status='ASSIGNED').count()
            new_workload.current_workload_score = new_workload.open_verifications_count + new_workload.open_tickets_count
            new_workload.save()
            
            Notification.objects.create(
                staff_recipient=staff,
                notification_type='WORK_ASSIGNMENT',
                title='Bulk Verification Tasks Assigned',
                message=f'You have been bulk-assigned {assigned_count} verification requests.',
                related_object_type='WorkAssignment',
                related_object_id=''
            )
            
        return ApiResponse(message=f'Successfully bulk assigned {assigned_count} verifications to {staff.get_full_name()}.')

class AdminQueueListView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    def get(self, request):
        from apps.core.models import SupportTicket, ProfileVerificationRequest
        
        queue_code = request.query_params.get('queue', 'UNASSIGNED')
        
        if queue_code == 'VERIFICATION':
            queryset = ProfileVerificationRequest.objects.all().select_related('member').order_by('-created_at')
            status_param = request.query_params.get('status')
            if status_param:
                queryset = queryset.filter(status=status_param)
            
            page = int(request.query_params.get('page', 1))
            page_size = int(request.query_params.get('page_size', 10))
            start = (page - 1) * page_size
            end = start + page_size
            
            total = queryset.count()
            results = queryset[start:end]
            
            data = [{
                'id': str(v.id),
                'member_name': v.member.get_full_name() if v.member else 'Unknown',
                'verification_type': v.verification_type,
                'verification_type_display': v.get_verification_type_display(),
                'status': v.status,
                'priority': v.priority,
                'submitted_at': v.submitted_at,
                'created_at': v.created_at,
            } for v in results]
            
            return ApiResponse(data={
                'count': total,
                'results': data
            })
        else:
            queryset = SupportTicket.objects.all().select_related('member', 'category').order_by('-created_at')
            
            if queue_code == 'UNASSIGNED':
                queryset = queryset.filter(status=SupportTicket.Status.UNASSIGNED)
            elif queue_code == 'SUPPORT':
                queryset = queryset.filter(status__in=[SupportTicket.Status.OPEN, SupportTicket.Status.ASSIGNED, SupportTicket.Status.IN_PROGRESS])
            elif queue_code == 'PAYMENT':
                queryset = queryset.filter(status__in=[SupportTicket.Status.OPEN, SupportTicket.Status.ASSIGNED, SupportTicket.Status.IN_PROGRESS], category__code__in=['PAYMENTS', 'REFUNDS', 'MEMBERSHIP'])
            elif queue_code == 'TECHNICAL':
                queryset = queryset.filter(status__in=[SupportTicket.Status.OPEN, SupportTicket.Status.ASSIGNED, SupportTicket.Status.IN_PROGRESS], category__code='TECHNICAL')
            elif queue_code == 'ESCALATED':
                queryset = queryset.filter(status=SupportTicket.Status.ESCALATED)
            elif queue_code == 'RESOLVED':
                queryset = queryset.filter(status=SupportTicket.Status.RESOLVED)
            elif queue_code == 'CLOSED':
                queryset = queryset.filter(status=SupportTicket.Status.CLOSED)
                
            category_param = request.query_params.get('category')
            if category_param:
                queryset = queryset.filter(category__id=category_param)
                
            priority_param = request.query_params.get('priority')
            if priority_param:
                queryset = queryset.filter(priority=priority_param)
                
            status_param = request.query_params.get('status')
            if status_param:
                queryset = queryset.filter(status=status_param)
                
            from apps.core.serializers import SupportTicketSerializer
            return paginated_response(request, queryset, SupportTicketSerializer)

class AdminAnalyticsDashboardView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    
    def get(self, request):
        from apps.core.models import SupportTicket, ProfileVerificationRequest, Queue, Workload
        from django.db.models import Avg, Sum
        from django.utils import timezone
        
        now = timezone.now()
        
        unassigned_count = SupportTicket.objects.filter(status=SupportTicket.Status.UNASSIGNED).count()
        open_tickets = SupportTicket.objects.filter(status__in=['OPEN', 'ASSIGNED', 'IN_PROGRESS']).count()
        escalated_count = SupportTicket.objects.filter(status=SupportTicket.Status.ESCALATED).count()
        resolved_today = SupportTicket.objects.filter(status=SupportTicket.Status.RESOLVED, resolved_at__date=now.date()).count()
        
        open_verifications = ProfileVerificationRequest.objects.filter(status__in=['PENDING', 'ASSIGNED', 'IN_REVIEW', 'RESUBMITTED']).count()
        escalated_verifications = ProfileVerificationRequest.objects.filter(status='ESCALATED').count()
        
        sla_violations = SupportTicket.objects.filter(
            status__in=['OPEN', 'ASSIGNED', 'IN_PROGRESS'],
            created_at__lt=now - timezone.timedelta(days=1)
        ).count()
        
        agent_stats = []

        department_performance = []
        from django.db.models import Count
        cat_counts = SupportTicket.objects.filter(status__in=['OPEN', 'ASSIGNED', 'IN_PROGRESS']).values('category__name').annotate(count=Count('id'))
        for item in cat_counts:
            department_performance.append({
                'name': item['category__name'],
                'open_count': item['count']
            })
            
        data = {
            'metrics': {
                'unassigned_tickets': unassigned_count,
                'open_tickets': open_tickets,
                'escalated_tickets': escalated_count,
                'resolved_tickets_today': resolved_today,
                'open_verifications': open_verifications,
                'escalated_verifications': escalated_verifications,
                'sla_violations': sla_violations,
                'avg_response_time': 45.5,
                'avg_resolution_time': 180.2,
            },
            'agent_performance': agent_stats,
            'department_performance': department_performance,
        }
        return ApiResponse(data=data)

class SpecializationViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = SpecializationSerializer
    queryset = Specialization.objects.all()

class QueueViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = QueueSerializer
    queryset = Queue.objects.all()

class AssignmentStrategyViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AssignmentStrategySerializer
    queryset = AssignmentStrategy.objects.all()

class EmployeeAvailabilityViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = EmployeeAvailabilitySerializer
    queryset = EmployeeAvailability.objects.all()

class WorkloadViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = WorkloadSerializer
    queryset = Workload.objects.all()

class AssignmentRuleViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AssignmentRuleSerializer
    queryset = AssignmentRule.objects.all()

class AssignmentAuditViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = AssignmentAuditSerializer
    queryset = AssignmentAudit.objects.all()

class SuperAdminMembershipPlanListCreateView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)

    def get(self, request):
        queryset = MembershipPlan.objects.all().order_by('display_order')
        return paginated_response(request, queryset, MembershipPlanSerializer)

    def post(self, request):
        serializer = MembershipPlanSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        price = serializer.validated_data.get('price')
        if price is not None and price < 0:
            return Response({'price': ['Price cannot be negative.']}, status=status.HTTP_400_BAD_REQUEST)
            
        slug = serializer.validated_data.get('slug')
        if slug and MembershipPlan.objects.filter(slug=slug).exists():
            return Response({'slug': ['Plan with this slug already exists.']}, status=status.HTTP_400_BAD_REQUEST)
            
        plan = serializer.save()
        audit(
            request, request.user, action='MEMBERSHIP_PLAN_CREATED', module='memberships',
            target_type='MembershipPlan', target_id=plan.pk,
        )
        return ApiResponse(data=MembershipPlanSerializer(plan).data, status=status.HTTP_201_CREATED)

class SuperAdminMembershipPlanDetailView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)

    def get(self, request, plan_id):
        plan = get_object_or_404(MembershipPlan, pk=plan_id)
        return ApiResponse(data=MembershipPlanSerializer(plan).data)

    def patch(self, request, plan_id):
        plan = get_object_or_404(MembershipPlan, pk=plan_id)
        previous_entitlements = dict(plan.entitlements or {})
        serializer = MembershipPlanSerializer(plan, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        
        price = serializer.validated_data.get('price')
        if price is not None and price < 0:
            return Response({'price': ['Price cannot be negative.']}, status=status.HTTP_400_BAD_REQUEST)
            
        slug = serializer.validated_data.get('slug')
        if slug and MembershipPlan.objects.filter(slug=slug).exclude(pk=plan_id).exists():
            return Response({'slug': ['Plan with this slug already exists.']}, status=status.HTTP_400_BAD_REQUEST)
            
        plan = serializer.save()
        audit(
            request, request.user, action='MEMBERSHIP_PLAN_UPDATED', module='memberships',
            target_type='MembershipPlan', target_id=plan.pk,
            old_data={'entitlements': previous_entitlements},
            new_data={'entitlements': plan.entitlements},
        )
        return ApiResponse(data=MembershipPlanSerializer(plan).data)

    def delete(self, request, plan_id):
        plan = get_object_or_404(MembershipPlan, pk=plan_id)
        
        from apps.core.models import Payment, MemberMembership
        if Payment.objects.filter(plan=plan).exists() or MemberMembership.objects.filter(plan=plan).exists():
            plan.is_active = False
            plan.save(update_fields=('is_active', 'updated_at'))
            audit(
                request, request.user, action='MEMBERSHIP_PLAN_ARCHIVED', module='memberships',
                target_type='MembershipPlan', target_id=plan.pk,
            )
            return ApiResponse(message='Plan cannot be hard deleted due to membership history. It has been deactivated/archived.')
        
        target_id = plan.pk
        plan.delete()
        audit(
            request, request.user, action='MEMBERSHIP_PLAN_DELETED', module='memberships',
            target_type='MembershipPlan', target_id=target_id,
        )
        return ApiResponse(message='Plan deleted.')

class SuperAdminMembershipPlanActivateView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)

    def post(self, request, plan_id):
        plan = get_object_or_404(MembershipPlan, pk=plan_id)
        plan.is_active = True
        plan.save(update_fields=('is_active', 'updated_at'))
        audit(
            request, request.user, action='MEMBERSHIP_PLAN_ACTIVATED', module='memberships',
            target_type='MembershipPlan', target_id=plan.pk,
        )
        return ApiResponse(data=MembershipPlanSerializer(plan).data, message='Membership plan activated.')

class SuperAdminMembershipPlanDeactivateView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN,)

    def post(self, request, plan_id):
        plan = get_object_or_404(MembershipPlan, pk=plan_id)
        plan.is_active = False
        plan.save(update_fields=('is_active', 'updated_at'))
        audit(
            request, request.user, action='MEMBERSHIP_PLAN_DEACTIVATED', module='memberships',
            target_type='MembershipPlan', target_id=plan.pk,
        )
        return ApiResponse(data=MembershipPlanSerializer(plan).data, message='Membership plan deactivated.')


class AdminReportExportView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)
    required_permission = 'reports.view'

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        report_type = request.query_params.get('report', 'users')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title='Report')

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        def style_header(ws, headers):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        def auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

        from django.utils import timezone
        from django.db.models import Count, Q
        from datetime import datetime

        now = timezone.now()
        if start_date:
            try:
                sd = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=now.tzinfo)
            except ValueError:
                sd = now - timedelta(days=30)
        else:
            sd = now - timedelta(days=30)

        if end_date:
            try:
                ed = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=now.tzinfo, hour=23, minute=59, second=59)
            except ValueError:
                ed = now
        else:
            ed = now

        if report_type == 'new_users':
            from apps.accounts.models import Member
            headers = ['ID', 'Name', 'Email', 'Mobile', 'Gender', 'Profile Status', 'Joined At']
            style_header(ws, headers)
            qs = Member.objects.filter(created_at__gte=sd, created_at__lte=ed).order_by('-created_at')
            for idx, m in enumerate(qs, 2):
                ws.cell(row=idx, column=1, value=str(m.id))
                ws.cell(row=idx, column=2, value=m.get_full_name() or m.email)
                ws.cell(row=idx, column=3, value=m.email)
                ws.cell(row=idx, column=4, value=str(m.mobile_number or ''))
                ws.cell(row=idx, column=5, value=getattr(m, 'gender', '') or '')
                ws.cell(row=idx, column=6, value=getattr(m, 'profile_status', '') or '')
                ws.cell(row=idx, column=7, value=m.created_at.strftime('%Y-%m-%d %H:%M') if m.created_at else '')
            ws.title = 'New Users'

        elif report_type == 'activities':
            from apps.accounts.models import SuperAdminActivityLog, AdminActivityLog, StaffActivityLog, MemberActivityLog, Member, Admin, Staff, SuperAdmin
            headers = ['Date', 'Actor ID', 'Actor Name', 'Role', 'Action', 'Module', 'Target Type', 'Target ID', 'Target Account', 'Description', 'IP Address']
            style_header(ws, headers)
            row_idx = 2
            for ModelCls, role in [
                (SuperAdminActivityLog, 'SUPER_ADMIN'),
                (AdminActivityLog, 'ADMIN'),
                (StaffActivityLog, 'ADMIN'),
                (MemberActivityLog, 'MEMBER'),
            ]:
                qs = ModelCls.objects.filter(created_at__gte=sd, created_at__lte=ed).order_by('-created_at')[:500]
                for log in qs:
                    target_info = ''
                    t_type = str(log.target_type or '').upper()
                    t_id = str(log.target_id or '')
                    if t_type in {'MEMBER', 'USER', 'MEMBERPROFILE'} and t_id:
                        m = Member.objects.filter(pk=t_id).first()
                        if m: target_info = f"{m.get_full_name()} ({m.email})"
                    elif t_type in {'ADMIN', 'STAFF', 'SUPER_ADMIN'} and t_id:
                        a = Admin.objects.filter(pk=t_id).first() or Staff.objects.filter(pk=t_id).first() or SuperAdmin.objects.filter(pk=t_id).first()
                        if a: target_info = f"{a.get_full_name()} ({a.email})"

                    ws.cell(row=row_idx, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M') if log.created_at else '')
                    ws.cell(row=row_idx, column=2, value=str(log.actor_id or ''))
                    ws.cell(row=row_idx, column=3, value=getattr(log, 'actor_name', '') or '')
                    ws.cell(row=row_idx, column=4, value=role)
                    ws.cell(row=row_idx, column=5, value=log.action or '')
                    ws.cell(row=row_idx, column=6, value=log.module or '')
                    ws.cell(row=row_idx, column=7, value=log.target_type or '')
                    ws.cell(row=row_idx, column=8, value=log.target_id or '')
                    ws.cell(row=row_idx, column=9, value=target_info)
                    ws.cell(row=row_idx, column=10, value=(log.description or '')[:200])
                    ws.cell(row=row_idx, column=11, value=log.ip_address or '')
                    row_idx += 1
            ws.title = 'Activities'

        elif report_type == 'expiring_memberships':
            from apps.core.services.membership_lifecycle_service import MembershipLifecycleService
            headers = ['Member ID', 'Name', 'Email', 'Mobile', 'Plan', 'Expires At', 'Days Remaining', 'Contact Status']
            style_header(ws, headers)
            members = MembershipLifecycleService.get_expiring_members_for_support(30)
            for idx, m in enumerate(members, 2):
                ws.cell(row=idx, column=1, value=m.get('member_id', ''))
                ws.cell(row=idx, column=2, value=m.get('member_name', ''))
                ws.cell(row=idx, column=3, value=m.get('email', ''))
                ws.cell(row=idx, column=4, value=m.get('mobile', ''))
                ws.cell(row=idx, column=5, value=m.get('plan', ''))
                ws.cell(row=idx, column=6, value=m.get('expires_at', ''))
                ws.cell(row=idx, column=7, value=m.get('days_remaining', ''))
                ws.cell(row=idx, column=8, value=m.get('contact_status', ''))
            ws.title = 'Expiring Memberships'

        elif report_type == 'tickets':
            from apps.core.models import SupportTicket
            from apps.core.serializers import TicketListSerializer
            headers = ['Ticket #', 'Subject', 'Member', 'Status', 'Priority', 'Category', 'Created At']
            style_header(ws, headers)
            qs = SupportTicket.objects.filter(created_at__gte=sd, created_at__lte=ed).select_related('member', 'category').order_by('-created_at')
            for idx, t in enumerate(qs, 2):
                ws.cell(row=idx, column=1, value=t.ticket_number or '')
                ws.cell(row=idx, column=2, value=t.subject or '')
                ws.cell(row=idx, column=3, value=t.member.email if t.member else '')
                ws.cell(row=idx, column=4, value=t.status or '')
                ws.cell(row=idx, column=5, value=t.priority or '')
                ws.cell(row=idx, column=6, value=t.category.name if t.category else '')
                ws.cell(row=idx, column=7, value=t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '')
            ws.title = 'Support Tickets'

        elif report_type == 'payments':
            from apps.core.models import Payment
            headers = ['ID', 'Member', 'Amount', 'Currency', 'Status', 'Gateway', 'Reference', 'Created At']
            style_header(ws, headers)
            qs = Payment.objects.filter(created_at__gte=sd, created_at__lte=ed).select_related('member').order_by('-created_at')
            for idx, p in enumerate(qs, 2):
                ws.cell(row=idx, column=1, value=str(p.id))
                ws.cell(row=idx, column=2, value=p.member.email if p.member else '')
                ws.cell(row=idx, column=3, value=str(p.amount or 0))
                ws.cell(row=idx, column=4, value=p.currency or 'INR')
                ws.cell(row=idx, column=5, value=p.status or '')
                ws.cell(row=idx, column=6, value=p.gateway or '')
                ws.cell(row=idx, column=7, value=p.gateway_reference or '')
                ws.cell(row=idx, column=8, value=p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else '')
            ws.title = 'Payments'

        else:
            headers = ['Report Type', 'Status']
            style_header(ws, headers)
            ws.cell(row=2, column=1, value=report_type)
            ws.cell(row=2, column=2, value='Unknown report type')

        auto_width(ws)

        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'{report_type}-report-{sd.strftime("%Y%m%d")}-to-{ed.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class AdminStaffPerformanceAnalyticsView(ScopedAPIView):
    allowed_account_types = (AccountType.SUPER_ADMIN, AccountType.ADMIN)

    def get(self, request):
        from datetime import timedelta
        from django.db.models import Q
        from django.utils import timezone
        from rest_framework.response import Response
        from apps.accounts.models import (
            Admin, AdminActivityLog, MemberDocument, StaffActivityLog,
        )
        from apps.profiles.models import ProfilePhoto
        from apps.core.models import SupportTicket

        start_str = request.query_params.get('start_date', '').strip()
        end_str = request.query_params.get('end_date', '').strip()
        staff_id = request.query_params.get('staff_id', '').strip()

        now = timezone.now()
        try:
            if start_str:
                start_date = timezone.datetime.strptime(start_str, '%Y-%m-%d').date()
            else:
                start_date = (now - timedelta(days=30)).date()
        except ValueError:
            start_date = (now - timedelta(days=30)).date()

        try:
            if end_str:
                end_date = timezone.datetime.strptime(end_str, '%Y-%m-%d').date()
            else:
                end_date = now.date()
        except ValueError:
            end_date = now.date()

        admin_qs = Admin.objects.filter(is_active=True)
        if str(request.user.account_type) != AccountType.SUPER_ADMIN:
            admin_qs = admin_qs.filter(pk=request.user.pk)
        elif staff_id and staff_id != 'all':
            admin_qs = admin_qs.filter(pk=staff_id)

        results = []
        days_count = (end_date - start_date).days + 1
        date_list = [start_date + timedelta(days=i) for i in range(days_count)]

        for staff_user in admin_qs:
            s_id = str(staff_user.pk)

            solved_tickets_qs = SupportTicket.objects.filter(
                Q(resolved_by_admin=staff_user) | Q(resolved_by_super_admin_id=s_id) | Q(resolver_id=staff_user.pk),
                resolved_at__date__gte=start_date,
                resolved_at__date__lte=end_date,
            )
            tickets_total = solved_tickets_qs.count()

            photos_qs = ProfilePhoto.objects.filter(
                Q(verified_by_admin_id=s_id) | Q(verified_by_super_admin_id=s_id) | Q(verified_by_staff_id=s_id),
                status=ProfilePhoto.Status.APPROVED,
                verified_at__date__gte=start_date,
                verified_at__date__lte=end_date,
            )
            photos_total = photos_qs.count()

            docs_qs = MemberDocument.objects.filter(
                Q(reviewed_by_id=staff_user.pk),
                status='APPROVED',
                uploaded_at__date__gte=start_date,
                uploaded_at__date__lte=end_date,
            )
            docs_total = docs_qs.count()

            activity_logs_qs = AdminActivityLog.objects.filter(
                actor_id=staff_user.pk,
                created_at__date__gte=start_date,
                created_at__date__lte=end_date,
            )

            vr_profiles_count = ProfileVerificationRequest.objects.filter(
                status=ProfileVerificationRequest.Status.APPROVED,
                verification_type=ProfileVerificationRequest.VerificationType.FULL_PROFILE,
            ).filter(
                Q(reviewed_by_admin_id=s_id) | Q(reviewed_by_super_admin_id=s_id) | Q(reviewed_by_staff_id=s_id),
                approved_at__date__gte=start_date,
                approved_at__date__lte=end_date,
            ).count()

            log_profiles_count = activity_logs_qs.filter(
                Q(action__icontains='APPROVE_PROFILE') | Q(action='MEMBER_APPROVE_PROFILE') | Q(action='PROFILE_APPROVED') | Q(description__icontains='approve profile') | Q(description__icontains='approved profile')
            ).count()

            profiles_total = max(vr_profiles_count, log_profiles_count)

            log_photos_count = activity_logs_qs.filter(
                Q(action__icontains='APPROVE_PHOTO') | Q(action='MEMBER_APPROVE_PHOTO') | Q(action='PHOTO_APPROVED') | Q(description__icontains='approve photo')
            ).count()
            photos_total = max(photos_qs.count(), log_photos_count)

            log_docs_count = activity_logs_qs.filter(
                Q(action__icontains='DOCUMENT') | Q(action='DOCUMENT_APPROVED') | Q(action='VERIFICATION_DOCUMENT_APPROVED') | Q(description__icontains='document_approved')
            ).count()
            docs_total = max(docs_qs.count(), log_docs_count)

            log_tickets_count = activity_logs_qs.filter(
                Q(action__icontains='TICKET_SOLVED') | Q(action='TICKET_RESOLVED')
            ).count()
            tickets_total = max(solved_tickets_qs.count(), log_tickets_count)

            daily_breakdown = []
            for dt in date_list:
                t_count = max(solved_tickets_qs.filter(resolved_at__date=dt).count(), activity_logs_qs.filter(Q(action__icontains='TICKET_SOLVED') | Q(action='TICKET_RESOLVED'), created_at__date=dt).count())
                p_count = max(photos_qs.filter(verified_at__date=dt).count(), activity_logs_qs.filter(Q(action__icontains='APPROVE_PHOTO') | Q(action='MEMBER_APPROVE_PHOTO') | Q(action='PHOTO_APPROVED'), created_at__date=dt).count())
                d_count = max(docs_qs.filter(uploaded_at__date=dt).count(), activity_logs_qs.filter(Q(action__icontains='DOCUMENT') | Q(action='DOCUMENT_APPROVED') | Q(action='VERIFICATION_DOCUMENT_APPROVED'), created_at__date=dt).count())
                prof_count = activity_logs_qs.filter(
                    Q(action__icontains='APPROVE_PROFILE') | Q(action='MEMBER_APPROVE_PROFILE') | Q(action='PROFILE_APPROVED') | Q(description__icontains='approve profile') | Q(description__icontains='approved profile'),
                    created_at__date=dt,
                ).count()
                daily_breakdown.append({
                    'date': dt.strftime('%Y-%m-%d'),
                    'tickets_solved': t_count,
                    'photos_approved': p_count,
                    'documents_approved': d_count,
                    'profiles_approved': prof_count,
                })

            results.append({
                'id': s_id,
                'full_name': staff_user.get_full_name() or staff_user.email,
                'email': staff_user.email,
                'mobile_number': getattr(staff_user, 'mobile_number', '') or '',
                'admin_id': getattr(staff_user, 'admin_id', '') or '',
                'photo': getattr(staff_user, 'photo', '') or '',
                'bio': getattr(staff_user, 'bio', '') or '',
                'role': staff_user.admin_role_code,
                'created_at': staff_user.created_at.strftime('%Y-%m-%d'),
                'summary': {
                    'tickets_solved': tickets_total,
                    'photos_approved': photos_total,
                    'documents_approved': docs_total,
                    'profiles_approved': profiles_total,
                },
                'daily_breakdown': daily_breakdown,
            })

        return Response({
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'staff_performance': results,
        })

